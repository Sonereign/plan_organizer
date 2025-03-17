import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

def get_headers(ws):
    """Returns a list of headers from the first row of the worksheet."""
    return [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

def find_last_category_column(headers):
    """Finds the last occurrence of the 'Category' column."""
    for col in range(len(headers), 0, -1):
        if headers[col - 1] == "Category":
            return col
    return None

def delete_columns_after_category(ws, last_category_index):
    """Deletes all columns after the last occurrence of 'Category'."""
    if last_category_index:
        ws.delete_cols(last_category_index + 1, ws.max_column - last_category_index)
        print(f"Deleted columns from index {last_category_index + 1} onward.")

def add_percentage_columns(ws, previous_years):
    """Adds 'Percent to Total YYYY' columns for the current year and given previous years."""
    headers = get_headers(ws)
    last_col = len(headers) + 1  # Insert after the last existing column
    current_year = datetime.now().year
    all_years = [current_year] + previous_years  # Include current year first

    for index, year in enumerate(all_years):
        ws.insert_cols(last_col + index)
        ws.cell(row=1, column=last_col + index).value = f"Percent to Total {year}"

    print(f"Added 'Percent to Total' columns for: {all_years}")

    return last_col + len(all_years)  # Return the next available column index

def find_and_replace_percent_to_total_column(ws, current_year):
    """
    Finds the first occurrence of the column 'Percent to Total {current_year}',
    replaces it with a black-filled separator, and returns the next available column index.
    """
    headers = get_headers(ws)
    target_header = f"Percent to Total {current_year}"

    for col in range(1, len(headers) + 1):
        if ws.cell(row=1, column=col).value == target_header:
            # Replace the header with an empty string
            ws.cell(row=1, column=col).value = ""

            # Fill the entire column with black
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col).fill = black_fill

            print(f"Replaced '{target_header}' column at position {col} with a black-filled separator.")
            return col + 1  # Return the next available column index

    print(f"Column '{target_header}' not found.")
    return None

def add_separator_column(ws, insert_at_col):
    """Adds a black-filled separator column at the specified position."""
    ws.insert_cols(insert_at_col)
    ws.cell(row=1, column=insert_at_col).value = ""  # Keeping the header empty

    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for row in range(1, ws.max_row + 1):  # Fill the entire column with black
        ws.cell(row=row, column=insert_at_col).fill = black_fill

    print(f"Added a black-filled separator column at position {insert_at_col}")

    return insert_at_col + 1  # Return the next available column index

def add_percent_difference_columns(ws, insert_at_col, previous_years):
    """Adds 'Percent difference current_year - previous_year' columns after the first separator."""
    current_year = datetime.now().year

    for index, prev_year in enumerate(previous_years):
        ws.insert_cols(insert_at_col + index)
        ws.cell(row=1, column=insert_at_col + index).value = f"Percent difference {current_year} - {prev_year}"

    print(f"Added 'Percent difference' columns for: {previous_years}")

    return insert_at_col + len(previous_years)  # Return the next available column index

def process_stage9(input_file, output_file, previous_years):
    """Processes Stage 9 by deleting columns after 'Category', adding percentage columns, separators, and percent differences."""
    print("#######################################################")
    print(f"Running Stage 9 with {input_file=} - {output_file=} - {previous_years=}")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    headers = get_headers(ws)

    """For some magic reason had to remove this.. ? who knows?"""
    # last_category_index = find_last_category_column(headers)
    # delete_columns_after_category(ws, last_category_index)

    next_available_col = add_percentage_columns(ws, previous_years)

    # Add the remaining columns and separators
    next_available_col = add_separator_column(ws, next_available_col)
    next_available_col = add_percent_difference_columns(ws, next_available_col, previous_years)
    add_separator_column(ws, next_available_col)  # Final black separator

    # Find and replace the 'Percent to Total {current_year}' column with a black-filled separator
    current_year = datetime.now().year
    find_and_replace_percent_to_total_column(ws, current_year)

    wb.save(output_file)
    print(f"Stage 9 processing complete. Output saved to {output_file}")

def stage9(input_path, output_path, previous_years):
    """Entry point for Stage 9 processing."""
    process_stage9(input_file=input_path, output_file=output_path, previous_years=previous_years)

if __name__ == '__main__':
    input_path = "stage8_output.xlsx"
    output_path = "stage9_output.xlsx"
    previous_years = [2024, 2023]  # Example: List of previous years
    stage9(input_path=input_path, output_path=output_path, previous_years=previous_years)