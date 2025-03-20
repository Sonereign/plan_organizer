from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from logger import logger

# Constants
GREEK_DAYS = {
    "Mon": "Δευ", "Tue": "Τρι", "Wed": "Τετ", "Thu": "Πεμ",
    "Fri": "Παρ", "Sat": "Σαβ", "Sun": "Κυρ"
}

DAY_COLORS = {
    "Παρ": "ADD8E6",  # Light Blue (Friday)
    "Σαβ": "90EE90",  # Light Green (Saturday)
    "Κυρ": "FFB6C1"  # Light Pink (Sunday)
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep"]
USE_FORMULAS = False
DO_CALCULATIONS = False


def load_and_prepare_data(input_file):
    """Load data from Excel and prepare it for processing."""
    try:
        logger.info(f"Loading data from {input_file}")
        df = pd.read_excel(input_file, header=None)
        headers = df.iloc[0]
        df = df[1:].reset_index(drop=True)
        df.columns = headers

        if "Capacity" in df.columns:
            logger.debug("Dropping 'Capacity' column")
            df = df.drop(columns=["Capacity"])

        logger.info("Data loaded successfully")
        return df, headers
    except Exception as e:
        logger.error(f"Error loading data from {input_file}: {e}", exc_info=True)
        raise


def format_date_column(col):
    """Format a single date column into Greek day and date."""
    try:
        date = pd.to_datetime(col, dayfirst=True, errors="coerce")
        if pd.notna(date):
            greek_day = GREEK_DAYS.get(date.strftime("%a"), date.strftime("%a"))
            formatted_date = f"{greek_day} {date.strftime('%d/%m')}"
            logger.debug(f"Formatted date '{col}' to '{formatted_date}'")
            return formatted_date
        else:
            logger.debug(f"Invalid date format for column: {col}")
            return col
    except Exception as e:
        logger.error(f"Error formatting date column '{col}': {e}", exc_info=True)
        return col


def format_dates(df):
    """Format all date columns in the DataFrame."""
    logger.info("Formatting date columns")
    try:
        date_columns = df.columns[1:]
        formatted_columns = [df.columns[0]] + [format_date_column(col) for col in date_columns]
        df.columns = formatted_columns
        logger.info("Date formatting completed")
        return df
    except Exception as e:
        logger.error(f"Error formatting date columns {e}", exc_info=True)
        raise


def find_camping_first_index(df):
    """Find the first row index where 'Camping' appears in the first column."""
    try:
        index = df[df.iloc[:, 0].astype(str).str.startswith("Camping")].index.min()
        if pd.isna(index):
            logger.warning("No 'Camping' section found in data")
            return None
        logger.info(f"'Camping' section starts at row index {index}")
        return index
    except Exception as e:
        logger.error("Error finding 'Camping' section index", exc_info=True)
        raise


def insert_totals_and_spacing(df, split_index):
    """Insert 'Total Rooms' before camping section and 'Total Camping' at the end, with one empty row in between."""
    try:
        if split_index is None:
            logger.warning("Skipping total insertion since no 'Camping' section found")
            return df

        logger.info("Inserting totals and spacing")
        total_rooms_row = pd.DataFrame([["Total Rooms"] + [""] * (len(df.columns) - 1)], columns=df.columns)
        total_camping_row = pd.DataFrame([["Total Camping"] + [""] * (len(df.columns) - 1)], columns=df.columns)
        empty_row = pd.DataFrame([[""] * len(df.columns)], columns=df.columns)

        # Split the dataframe into room and camping sections
        top_part = df.iloc[:split_index]
        bottom_part = df.iloc[split_index:]

        # Concatenate parts
        df = pd.concat([top_part, total_rooms_row, empty_row, bottom_part, total_camping_row], ignore_index=True)

        logger.info("Totals and spacing inserted successfully")
        return df
    except Exception as e:
        logger.error(f"Error inserting totals and spacing {e}", exc_info=True)
        raise


def apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col):
    try:
        if not USE_FORMULAS:
            logger.debug("USE_FORMULAS is False, applying direct sum calculations")
            return apply_column_sums_noform(ws, total_rooms_row, total_camping_row, max_col)

        if DO_CALCULATIONS:
            logger.info("Applying Excel sum formulas")
            for col in range(2, max_col + 2):
                col_letter = get_column_letter(col)
                if total_rooms_row:
                    ws.cell(row=total_rooms_row,
                            column=col).value = f"=SUM({col_letter}2:{col_letter}{total_rooms_row - 1})"
                if total_camping_row:
                    ws.cell(row=total_camping_row,
                            column=col).value = f"=SUM({col_letter}{total_rooms_row + 2}:{col_letter}{total_camping_row - 2})"
            logger.info("Sum formulas applied successfully")
    except Exception as e:
        logger.error("Error applying sum formulas", exc_info=True)
        raise


def apply_column_sums_noform(ws, total_rooms_row, total_camping_row, max_col):
    """Directly calculate and insert column sums without using Excel formulas."""
    try:
        if DO_CALCULATIONS:
            logger.info("Calculating column sums without formulas")
            for col in range(2, max_col + 2):
                column_sum_rooms = 0
                column_sum_camping = 0

                # Calculate sum for rooms
                if total_rooms_row:
                    for row in range(2, total_rooms_row):
                        value = ws.cell(row=row, column=col).value
                        if isinstance(value, (int, float)):
                            column_sum_rooms += value
                    ws.cell(row=total_rooms_row, column=col).value = column_sum_rooms
                    logger.debug(f"Column {col} rooms sum: {column_sum_rooms}")

                # Calculate sum for camping
                if total_camping_row:
                    for row in range(total_rooms_row + 2, total_camping_row):
                        value = ws.cell(row=row, column=col).value
                        if isinstance(value, (int, float)):
                            column_sum_camping += value
                    ws.cell(row=total_camping_row, column=col).value = column_sum_camping
                    logger.debug(f"Column {col} camping sum: {column_sum_camping}")

            logger.info("Direct column sums applied successfully")
    except Exception as e:
        logger.error("Error applying direct column sums", exc_info=True)
        raise


def apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row):
    """Apply Excel formulas to calculate row sums and percentages."""
    try:
        logger.info("Applying row sum formulas")
        total_column = max_col + 1
        percent_column = total_column + 1
        current_year = datetime.now().year

        add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year)
        add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row,
                              current_year)
        add_monthly_sums(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row)

        logger.info("Row sum formulas applied successfully")
    except Exception as e:
        logger.error("Error applying row sum formulas", exc_info=True)
        raise


def add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year):
    """Add a 'Total' column to calculate row sums."""
    try:
        if not USE_FORMULAS:
            return add_total_column_direct_noform(ws, max_row, max_col, total_column, total_rooms_row,
                                                  total_camping_row, current_year)
        logger.info("Adding total column")
        ws.cell(row=1, column=total_column).value = f"Total {current_year}"
        ws.cell(row=1, column=total_column).font = Font(bold=True)

        if DO_CALCULATIONS:
            for row in range(2, max_row + 1):
                if row not in [total_rooms_row, total_camping_row]:
                    first_col_letter = ws.cell(row=row, column=2).column_letter
                    last_col_letter = ws.cell(row=row, column=max_col).column_letter
                    ws.cell(row=row,
                            column=total_column).value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                    ws.cell(row=row, column=total_column).fill = YELLOW_FILL
                    ws.cell(row=row, column=total_column).font = Font(bold=True)
                    logger.debug(f"Applied SUM formula for row {row}")

        logger.info("Total column added successfully")
    except Exception as e:
        logger.error(f"Error adding total column {e}", exc_info=True)
        raise


def add_total_column_direct_noform(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row,
                                   current_year):
    """Directly calculate and insert row sums into the 'Total' column without using Excel formulas."""
    try:
        logger.info("Adding total column without formulas")
        ws.cell(row=1, column=total_column).value = f"Total {current_year}"
        ws.cell(row=1, column=total_column).font = Font(bold=True)

        if DO_CALCULATIONS:
            for row in range(2, max_row + 1):
                if row not in [total_rooms_row, total_camping_row]:
                    row_sum = 0
                    for col in range(2, max_col + 1):
                        value = ws.cell(row=row, column=col).value
                        if isinstance(value, (int, float)):
                            row_sum += value

                    ws.cell(row=row, column=total_column).value = row_sum
                    ws.cell(row=row, column=total_column).fill = YELLOW_FILL
                    ws.cell(row=row, column=total_column).font = Font(bold=True)
                    logger.debug(f"Computed row sum for row {row}: {row_sum}")

        logger.info("Total column (direct calculation) added successfully")
    except Exception as e:
        logger.error(f"Error adding total column (direct calculation) {e}", exc_info=True)
        raise


def add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, current_year):
    """Add a 'Percent to Total' column to calculate percentages."""
    try:
        if not USE_FORMULAS:
            return add_percentage_column_direct_noform(ws, max_row, total_column, percent_column, total_rooms_row,
                                                       total_camping_row, current_year)
        logger.info("Adding percentage column")
        ws.cell(row=1, column=percent_column).value = f"Percent to Total {current_year}"
        ws.cell(row=1, column=percent_column).font = Font(bold=True)
        if DO_CALCULATIONS:
            for row in range(2, max_row + 1):
                if row not in [total_rooms_row, total_camping_row]:
                    total_rooms_col_letter = ws.cell(row=total_rooms_row, column=total_column).column_letter
                    total_camping_col_letter = ws.cell(row=total_camping_row, column=total_column).column_letter
                    current_row_col_letter = ws.cell(row=row, column=total_column).column_letter

                    if row < total_rooms_row:
                        ws.cell(row=row,
                                column=percent_column).value = f"=IF({total_rooms_col_letter}{total_rooms_row}<>0, {current_row_col_letter}{row}/{total_rooms_col_letter}{total_rooms_row}, 0)"
                    elif row > total_rooms_row + 1:
                        ws.cell(row=row,
                                column=percent_column).value = f"=IF({total_camping_col_letter}{total_camping_row}<>0, {current_row_col_letter}{row}/{total_camping_col_letter}{total_camping_row}, 0)"

                    ws.cell(row=row, column=percent_column).number_format = "0.00%"
                    logger.debug(f"Applied percentage formula for row {row}")

        logger.info("Percentage column added successfully")
    except Exception as e:
        logger.error("Error adding percentage column", exc_info=True)
        raise


def add_percentage_column_direct_noform(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row,
                                        current_year):
    """Add a 'Percent to Total' column and calculate percentages directly."""
    try:
        logger.info("Adding percentage column without formulas")
        ws.cell(row=1, column=percent_column).value = f"Percent to Total {current_year}"
        ws.cell(row=1, column=percent_column).font = Font(bold=True)

        total_rooms_value = 0
        total_camping_value = 0

        for row in range(2, total_rooms_row):
            value = ws.cell(row=row, column=total_column).value
            if isinstance(value, (int, float)):
                total_rooms_value += value

        for row in range(total_rooms_row + 2, max_row + 1):
            value = ws.cell(row=row, column=total_column).value
            if isinstance(value, (int, float)):
                total_camping_value += value

        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                value = ws.cell(row=row, column=total_column).value
                if isinstance(value, (int, float)):
                    if row < total_rooms_row:
                        percentage = value / total_rooms_value if total_rooms_value else 0
                    elif row > total_rooms_row + 1:
                        percentage = value / total_camping_value if total_camping_value else 0
                    else:
                        percentage = 0

                    ws.cell(row=row, column=percent_column).value = percentage
                    ws.cell(row=row, column=percent_column).number_format = "0.00%"
                    logger.debug(f"Computed percentage for row {row}: {percentage:.2%}")

        logger.info("Percentage column (direct calculation) added successfully")
    except Exception as e:
        logger.error(f"Error adding percentage column (direct calculation) {e}", exc_info=True)
        raise


def add_monthly_sums(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row):
    if not USE_FORMULAS:
        logger.info("Skipping formulas. Using direct calculations for monthly sums.")
        return add_monthly_sums_direct_noform(ws, max_row, total_column, separator_column_2, total_rooms_row,
                                              total_camping_row)

    logger.info("Adding monthly sum columns with formulas.")
    try:
        month_ranges = find_monthly_column_ranges(ws, total_column)
        month_start_col = separator_column_2 + 1  # Start after the second separator

        for i, month in enumerate(MONTHS):
            month_col = month_start_col + i
            ws.cell(row=1, column=month_col).value = f"{month} 2025"
            ws.cell(row=1, column=month_col).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=month_col).column_letter].width = 12
            logger.debug(f"Added header for {month} 2025 at column {month_col}.")

            if month in month_ranges:
                first_col_letter = ws.cell(row=1, column=month_ranges[month][0]).column_letter
                last_col_letter = ws.cell(row=1, column=month_ranges[month][1]).column_letter
                logger.debug(f"{month} range: {first_col_letter} to {last_col_letter}.")

                if DO_CALCULATIONS:
                    for row in range(2, max_row + 1):
                        if row not in [total_rooms_row, total_camping_row]:
                            ws.cell(row=row,
                                    column=month_col).value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                            logger.debug(f"Set formula for row {row}, column {month_col}.")

    except Exception as e:
        logger.error(f"Error in add_monthly_sums: {e}", exc_info=True)


def add_monthly_sums_direct_noform(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row):
    logger.info("Adding monthly sums directly without formulas.")

    try:
        month_ranges = find_monthly_column_ranges(ws, total_column)
        month_start_col = separator_column_2 + 1  # Start after the second separator

        for i, month in enumerate(MONTHS):
            month_col = month_start_col + i
            ws.cell(row=1, column=month_col).value = f"{month} 2025"
            ws.cell(row=1, column=month_col).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=month_col).column_letter].width = 12
            logger.debug(f"Added header for {month} 2025 at column {month_col}.")

            if DO_CALCULATIONS:
                if month in month_ranges:
                    first_col_index = month_ranges[month][0] - 1  # Adjust for 0-based indexing
                    last_col_index = month_ranges[month][1] - 1  # Adjust for 0-based indexing

                    for row in range(2, max_row + 1):
                        if row not in [total_rooms_row, total_camping_row]:
                            row_sum = 0
                            for col in range(first_col_index, last_col_index + 1):
                                cell_value = ws.cell(row=row, column=col + 1).value  # Adjust for 1-based indexing
                                if isinstance(cell_value, (int, float)):
                                    row_sum += cell_value

                            ws.cell(row=row, column=month_col).value = row_sum
                            logger.debug(f"Row {row}, column {month_col} sum: {row_sum}")

    except Exception as e:
        logger.error(f"Error in add_monthly_sums_direct_noform: {e}", exc_info=True)


def find_monthly_column_ranges(ws, total_column):
    """Find the first and last column for each month."""
    logger.info("Finding monthly column ranges.")

    month_ranges = {}
    try:
        for col in range(2, total_column):
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                for i, month in enumerate(MONTHS):
                    month_number = str(i + 4).zfill(2)  # "Apr" = 04, "May" = 05, etc.
                    if header_value.endswith(f'/{month_number}'):
                        if month not in month_ranges:
                            month_ranges[month] = [col, col]
                        else:
                            month_ranges[month][1] = col
        logger.debug(f"Monthly column ranges: {month_ranges}")
    except Exception as e:
        logger.error(f"Error in find_monthly_column_ranges: {e}", exc_info=True)

    return month_ranges


def add_separator_column(ws, max_row, separator_column):
    """Add a black-filled separator column."""
    logger.info(f"Adding separator column at index {separator_column}.")

    try:
        for row in range(1, max_row + 1):
            separator_cell = ws.cell(row=row, column=separator_column)
            separator_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws.column_dimensions[ws.cell(row=1, column=separator_column).column_letter].width = 3
        logger.debug("Separator column added successfully.")
    except Exception as e:
        logger.error(f"Error in add_separator_column: {e}", exc_info=True)


def apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row):
    """Apply formatting to the worksheet."""
    logger.info("Applying formatting to the worksheet.")

    try:
        # Bold headers and center alignment
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        logger.debug("Applied bold font and center alignment to header row.")

        # Color day columns based on predefined colors
        for col_num in range(2, max_col + 1):
            column_title = ws.cell(row=1, column=col_num).value
            if column_title:
                day_part = column_title.split(" ")[0]
                if day_part in DAY_COLORS:
                    ws.cell(row=1, column=col_num).fill = PatternFill(
                        start_color=DAY_COLORS[day_part], end_color=DAY_COLORS[day_part], fill_type="solid"
                    )
                    logger.debug(f"Colored column {col_num} ({column_title}) with {DAY_COLORS[day_part]}.")

        # Highlight total rows
        for row in [total_rooms_row, total_camping_row]:
            if row:
                for col in range(1, max_col + 2):
                    ws.cell(row=row, column=col).fill = YELLOW_FILL
                    ws.cell(row=row, column=col).font = Font(bold=True)
                logger.debug(f"Formatted total row at index {row}.")

        # Apply border to all cells
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col + 1):
            for cell in row:
                cell.border = THIN_BORDER

        # Freeze top-left section
        ws.freeze_panes = "B2"
        logger.info("Formatting applied successfully.")
    except Exception as e:
        logger.error(f"Error in apply_formatting: {e}", exc_info=True)


def apply_excel_formatting_and_formulas(output_file):
    """Apply formatting and formulas to the output Excel file."""
    logger.info(f"Applying formatting and formulas to {output_file}.")

    try:
        wb = load_workbook(output_file)
        ws = wb.active
        max_col = ws.max_column
        max_row = ws.max_row
        total_rooms_row = total_camping_row = None

        # Identify total rows
        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "Total Rooms":
                total_rooms_row = row
            elif cell_value == "Total Camping":
                total_camping_row = row

        logger.debug(f"Identified total rows - Total Rooms: {total_rooms_row}, Total Camping: {total_camping_row}.")

        # Apply calculations and formatting
        apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col)
        apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row)
        apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row)

        wb.save(output_file)
        logger.info(f"Formatting and formulas applied successfully. File saved as {output_file}.")

    except Exception as e:
        logger.error(f"Error in apply_excel_formatting_and_formulas: {e}", exc_info=True)


def per_nationality_stage1(input_file, output_file):
    """Process reservations and generate the output Excel file."""
    logger.info("#######################################################")
    logger.info(f"Running Stage 5 with input file: {input_file}.")

    try:
        df, headers = load_and_prepare_data(input_file)
        df = format_dates(df)
        split_index = find_camping_first_index(df)
        df = insert_totals_and_spacing(df, split_index)

        df.to_excel(output_file, index=False, engine='openpyxl')
        logger.info(f"Data processing completed. Output saved to {output_file}.")

        apply_excel_formatting_and_formulas(output_file)
        logger.info(f"Stage 5 completed. Final output file: {output_file}.")

    except Exception as e:
        logger.error(f"Error in per_nationality_stage1: {e}", exc_info=True)


if __name__ == "__main__":
    # Default file paths (for standalone execution)
    INPUT_FILE = "sources/availabilityPerNationality2025.xls"
    OUTPUT_FILE = "nat_stage1_output.xlsx"

    # Run stage5
    per_nationality_stage1(INPUT_FILE, OUTPUT_FILE)
