import openpyxl
import re
from datetime import datetime


def extract_date(header):
    """Extracts datetime object from headers like 'Apr 2025'."""
    if not isinstance(header, str):
        return None
    try:
        return datetime.strptime(header, '%b %Y')
    except ValueError:
        return None


def get_headers(ws):
    """Returns a list of headers from the worksheet."""
    return [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]


def identify_date_columns(headers, current_year):
    """Identifies columns with 'Month Year' format for the current year."""
    return [(col, h, extract_date(h)) for col, h in enumerate(headers, start=1) if
            extract_date(h) and extract_date(h).year == current_year]


def insert_empty_columns(ws, date_columns, number_of_previous_year_data):
    """Inserts empty columns after each identified date column."""
    for offset, (col, header, _) in enumerate(date_columns):
        insert_col = col + offset * number_of_previous_year_data + 1
        for _ in range(number_of_previous_year_data):
            ws.insert_cols(insert_col)


def move_previous_years_data(ws, number_of_previous_year_data, current_year):
    """Moves previous years' data into the created empty columns and deletes the original columns."""
    headers = get_headers(ws)

    for month in range(1, 13):
        month_name = datetime(current_year, month, 1).strftime('%b')
        latest_header = f"{month_name} {current_year}"
        if latest_header not in headers:
            continue

        latest_col = headers.index(latest_header) + 1

        for year_offset in range(1, number_of_previous_year_data + 1):
            prev_year = current_year - year_offset
            prev_header = f"{month_name} {prev_year}"
            if prev_header not in headers:
                continue

            prev_col = headers.index(prev_header) + 1
            ws.move_range(
                f"{openpyxl.utils.get_column_letter(prev_col)}1:{openpyxl.utils.get_column_letter(prev_col)}{ws.max_row}",
                cols=(latest_col + year_offset - prev_col))
            ws.delete_cols(prev_col)
            headers = get_headers(ws)


def find_total_columns(ws):
    """Finds all 'Total YYYY' columns after the second occurrence of 'Category'."""
    category_count = 0
    total_columns = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value == "Category":
            category_count += 1
            continue
        if category_count == 2 and isinstance(cell_value, str) and re.match(r"Total \d{4}", cell_value):
            total_columns.append((col, cell_value))
    return total_columns


def insert_total_columns(ws, total_columns, number_of_previous_year_data, current_year, previous_years):
    """Repositions existing 'Total YYYY' columns instead of inserting new ones."""
    if not total_columns:
        return

    first_total_col = total_columns[0][0]

    # Insert a new column for "Total {current_year}"
    ws.insert_cols(first_total_col)
    ws.cell(row=1, column=first_total_col).value = f"Total {current_year}"

    # Move existing "Total YYYY" columns for previous years
    extra_index = 2
    previous_year_index = 1
    for _ in range(number_of_previous_year_data - 1):
        previous_total_header = f"Total {previous_years[previous_year_index]}"
        headers = get_headers(ws)

        if previous_total_header in headers:
            prev_col = headers.index(previous_total_header) + 1
            ws.move_range(
                f"{openpyxl.utils.get_column_letter(prev_col)}1:"
                f"{openpyxl.utils.get_column_letter(prev_col)}{ws.max_row}",
                cols=(first_total_col + extra_index - prev_col)
            )
            headers = get_headers(ws)  # Update headers after movement
        else:
            print(f"Warning: {previous_total_header} not found, skipping.")

        previous_year_index += 1
        extra_index += 1



def drop_total_current_year_column(ws, current_year):
    """Drops the first occurrence of the column with header 'Total current_year'."""
    headers = get_headers(ws)
    total_header = f"Total {current_year}"
    if total_header in headers:
        total_col = headers.index(total_header) + 1
        ws.delete_cols(total_col)
        print(f"Dropped first occurrence of column: {total_header} at position {total_col}")
    else:
        print(f"Column {total_header} not found, no deletion performed.")


def process_stage8(stage7_file, output_file, number_of_previous_year_data, previous_years):
    """Processes Stage 8 by inserting empty columns, moving data, and adjusting total columns."""
    wb = openpyxl.load_workbook(stage7_file)
    ws = wb.active
    current_year = datetime.now().year

    headers = get_headers(ws)
    date_columns = identify_date_columns(headers, current_year)
    insert_empty_columns(ws, date_columns, number_of_previous_year_data)
    move_previous_years_data(ws, number_of_previous_year_data, current_year)
    total_columns = find_total_columns(ws)
    insert_total_columns(ws, total_columns, number_of_previous_year_data, current_year, previous_years)
    drop_total_current_year_column(ws, current_year)

    wb.save(output_file)


def stage8(stage7_path, output_path, previous_years, number_of_previous_year_data):
    """Entry point for Stage 8 processing."""
    process_stage8(stage7_file=stage7_path, output_file=output_path, number_of_previous_year_data=number_of_previous_year_data, previous_years=previous_years)


if __name__ == '__main__':
    stage7_path = "stage7_output.xlsx"
    output_path = "stage8_output.xlsx"
    previous_years = ["2024", "2023"]  # List of previous years as strings
    number_of_previous_year_data = len(previous_years)
    stage8(stage7_path=stage7_path, output_path=output_path, previous_years=previous_years, number_of_previous_year_data=number_of_previous_year_data)
