import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
HOUSE_KEYWORDS = ["Beach Apt", ".LUX for 4", ".Safari Tent 5pax", ".Sea Safari 4pax",
                  ".Skyline 3pax", ".Standard Mobile Home", ".ΤΡΟΧΟΣΠΙΤΑ DELUXE",
                  ".ΤΡΟΧΟΣΠΙΤΑ SEA VIEW", ".ΤΡΟΧΟΣΠΙΤΑ standard"]
YOUTH_HOSTEL_KEYWORDS = ["Youth Hostel"]

DAY_COLORS = {
    "Fri": "ADD8E6",  # Light Blue (Friday)
    "Sat": "90EE90",  # Light Green (Saturday)
    "Sun": "FFB6C1"   # Light Pink (Sunday)
}


def load_data(input_file):
    """Load Excel file and extract first sheet."""
    df = pd.read_excel(input_file, sheet_name=None)
    sheet_name = list(df.keys())[0]  # Get first sheet
    df = df[sheet_name]
    df.iloc[:, 0] = df.iloc[:, 0].astype(str)  # Ensure Category column is a string
    return df


def detect_date_columns(df):
    """Detect first and last date columns."""
    date_cols = [col for col in df.columns if is_date(col)]
    if date_cols:
        return date_cols[0], date_cols[-1]  # Return first and last date column
    return None, None


def is_date(column_name):
    """Check if a column name is a date."""
    try:
        pd.to_datetime(column_name, dayfirst=True)
        return True
    except (ValueError, TypeError):
        return False


def format_date_columns(df, first_date_col, last_date_col):
    """Format date columns from first to last to 'Mon 14/9'."""
    if first_date_col and last_date_col:
        date_range = df.loc[:, first_date_col:last_date_col].columns
        df.rename(columns={col: pd.to_datetime(col, dayfirst=True).strftime("%a %d/%m") for col in date_range},
                  inplace=True)


def contains_keyword(value, keywords):
    """Check if a value contains any keyword from a list."""
    return any(keyword.lower() in value.lower() for keyword in keywords)


def add_empty_separator(df):
    """Create an empty row separator matching the number of columns in df."""
    return pd.DataFrame([[""] * df.shape[1]], columns=df.columns)


def split_sections(df):
    """Split the dataframe into Houses, Youth Hostel, and The Rest."""
    houses = df[df.iloc[:, 0].apply(lambda x: contains_keyword(x, HOUSE_KEYWORDS))]
    youth_hostel = df[df.iloc[:, 0].apply(lambda x: contains_keyword(x, YOUTH_HOSTEL_KEYWORDS))]
    the_rest = df[~df.index.isin(houses.index) & ~df.index.isin(youth_hostel.index)]

    # Add empty row separators
    return pd.concat([houses, add_empty_separator(df), youth_hostel, add_empty_separator(df), the_rest],
                     ignore_index=True)


def save_to_excel(df, output_file):
    """Save dataframe to an Excel file."""
    df.to_excel(output_file, index=False, engine='openpyxl')


def apply_day_colors(output_file):
    """Apply colors only to the date header cells based on DAY_COLORS."""
    wb = load_workbook(output_file)
    ws = wb.active  # Get the active sheet

    # Get header row
    headers = [cell.value for cell in ws[1]]

    for col_idx, col_name in enumerate(headers, start=1):
        if isinstance(col_name, str) and len(col_name) > 3:  # Check if formatted as 'Mon 14/9'
            day = col_name[:3]  # Extract day part (e.g., "Mon")
            if day in DAY_COLORS:
                fill = PatternFill(start_color=DAY_COLORS[day], end_color=DAY_COLORS[day], fill_type="solid")

                # Apply color **only to the header row**
                ws.cell(row=1, column=col_idx).fill = fill

    wb.save(output_file)


def stage1(input_file, output_file):
    """
    Process the input file (availabilityPerZone) and save the result to the output file.
    """
    df = load_data(input_file)
    first_date_col, last_date_col = detect_date_columns(df)

    if first_date_col and last_date_col:
        format_date_columns(df, first_date_col, last_date_col)

    df_split = split_sections(df)
    save_to_excel(df_split, output_file)
    apply_day_colors(output_file)

    print(f"Stage 1 completed. File saved as {output_file}")


if __name__ == "__main__":
    # Default file paths (for standalone execution)
    INPUT_FILE = "availabilityPerZone2025.xls"
    OUTPUT_FILE = "stage1_output.xlsx"

    # Run stage1
    stage1(INPUT_FILE, OUTPUT_FILE)