import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from logger import logger

DO_CALCULATIONS = False

# Constants
GREEK_DAYS = {
    "Mon": "Δευ", "Tue": "Τρι", "Wed": "Τετ", "Thu": "Πεμ",
    "Fri": "Παρ", "Sat": "Σαβ", "Sun": "Κυρ"
}

DAY_COLORS = {
    "Παρ": "ADD8E6",  # Light Blue (Friday)
    "Σαβ": "90EE90",  # Light Green (Saturday)
    "Κυρ": "FFB6C1"  # Light Pink (Sunday)
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep"]


def load_and_prepare_data(input_file):
    """Load data from Excel and prepare it for processing."""
    df = pd.read_excel(input_file, header=None)
    headers = df.iloc[0]
    df = df[1:].reset_index(drop=True)
    df.columns = headers
    if "Capacity" in df.columns:
        df = df.drop(columns=["Capacity"])
    return df, headers


def format_date_column(col):
    """Format a single date column into Greek day and date."""
    try:
        date = pd.to_datetime(col, dayfirst=True, errors="coerce")
        if pd.notna(date):
            greek_day = GREEK_DAYS.get(date.strftime("%a"), date.strftime("%a"))
            return f"{greek_day} {date.strftime('%d/%m')}"
        else:
            return col
    except Exception:
        return col


def format_dates(df):
    """Format all date columns in the DataFrame."""
    date_columns = df.columns[1:]
    formatted_columns = [df.columns[0]] + [format_date_column(col) for col in date_columns]
    df.columns = formatted_columns
    return df


def find_camping_first_index(df):
    """Find the first row index where 'Camping' appears in the first column."""
    return df[df.iloc[:, 0].astype(str).str.startswith("Camping")].index.min()


def insert_totals_and_spacing(df, split_index, year):
    """Insert 'Total Rooms' before camping section and 'Total Camping' at the end, with one empty row in between."""
    total_rooms_row = pd.DataFrame([["Total Rooms"] + [""] * (len(df.columns) - 1)], columns=df.columns)
    total_camping_row = pd.DataFrame([["Total Camping"] + [""] * (len(df.columns) - 1)], columns=df.columns)
    empty_row = pd.DataFrame([[""] * len(df.columns)], columns=df.columns)

    # Split the dataframe into room and camping sections
    top_part = df.iloc[:split_index]
    bottom_part = df.iloc[split_index:]

    # Concatenate parts, ensuring correct order and no extra empty row at the end
    df = pd.concat([top_part, total_rooms_row, empty_row, bottom_part, total_camping_row], ignore_index=True)

    return df


def apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col):
    if DO_CALCULATIONS:
        """Calculate column sums directly and insert the values instead of formulas."""
        for col in range(2, max_col + 2):
            column_values = [
                ws.cell(row=row, column=col).value for row in range(2, total_rooms_row)
                if isinstance(ws.cell(row=row, column=col).value, (int, float))
            ]

            if total_rooms_row:
                ws.cell(row=total_rooms_row, column=col).value = sum(column_values)

            column_values_camping = [
                ws.cell(row=row, column=col).value for row in range(total_rooms_row + 2, total_camping_row)
                if isinstance(ws.cell(row=row, column=col).value, (int, float))
            ]

            if total_camping_row:
                ws.cell(row=total_camping_row, column=col).value = sum(column_values_camping)


def apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row, year):
    """Apply Excel formulas to calculate row sums and percentages."""
    total_column = max_col + 1
    percent_column = total_column + 1  # "Percent to Total" column

    # Add the "Total" column
    add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, year=year)

    # Add the "Percent to Total" column
    # add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, year=year)

    add_monthly_sums(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, year=year)


def add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, year):
    """Add a 'Total' column and calculate row sums directly in Python."""
    ws.cell(row=1, column=total_column).value = f"Total {year}"
    ws.cell(row=1, column=total_column).font = Font(bold=True)

    for row in range(2, max_row + 1):
        # Compute sum of numeric values in the row
        row_sum = sum(
            ws.cell(row=row, column=col).value or 0
            for col in range(2, max_col + 1)
            if isinstance(ws.cell(row=row, column=col).value, (int, float))
        )
        ws.cell(row=row, column=total_column).value = row_sum
        ws.cell(row=row, column=total_column).fill = YELLOW_FILL
        ws.cell(row=row, column=total_column).font = Font(bold=True)


def add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, year):
    """Add a 'Percent to Total' column with static percentage values instead of formulas."""
    ws.cell(row=1, column=percent_column).value = f"Percent to Total {year}"
    ws.cell(row=1, column=percent_column).font = Font(bold=True)

    if DO_CALCULATIONS:
        total_rooms_value = ws.cell(row=total_rooms_row, column=total_column).value
        total_camping_value = ws.cell(row=total_camping_row, column=total_column).value

        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                current_value = ws.cell(row=row, column=total_column).value

                if row < total_rooms_row and total_rooms_value and current_value:
                    percentage = (current_value / total_rooms_value) * 100
                elif row > total_rooms_row + 1 and total_camping_value and current_value:
                    percentage = (current_value / total_camping_value) * 100
                else:
                    percentage = 0

                ws.cell(row=row, column=percent_column).value = round(percentage, 2) / 100  # Convert to decimal format
                ws.cell(row=row, column=percent_column).number_format = "0.00%"

        ws.column_dimensions[ws.cell(row=1, column=percent_column).column_letter].width = 15


def add_monthly_sums(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row, year):
    """Add monthly sum columns and calculate their sums directly in Python."""
    month_ranges = find_monthly_column_ranges(ws, total_column)
    month_start_col = separator_column_2 + 1  # Start after the second separator

    for i, month in enumerate(MONTHS):
        month_col = month_start_col + i
        ws.cell(row=1, column=month_col).value = f"{month} {year}"  # Include year in header
        ws.cell(row=1, column=month_col).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=month_col).column_letter].width = 12

        if month in month_ranges:
            start_col, end_col = month_ranges[month]

            for row in range(2, max_row + 1):
                if row not in [total_rooms_row, total_camping_row]:
                    # Compute sum of numeric values in the row for the given month's range
                    values = [
                        ws.cell(row=row, column=col).value or 0
                        for col in range(start_col, end_col + 1)
                        if isinstance(ws.cell(row=row, column=col).value, (int, float))
                    ]
                    monthly_sum = sum(values)

                    ws.cell(row=row, column=month_col).value = monthly_sum


def find_monthly_column_ranges(ws, total_column):
    """Find the first and last column for each month."""
    month_ranges = {}
    for col in range(2, total_column):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            for i, month in enumerate(MONTHS):
                month_number = str(i + 4).zfill(2)  # "Apr" = 04, "May" = 05, etc.
                if header_value.endswith(f'/{month_number}'):
                    if month not in month_ranges:
                        month_ranges[month] = [col, col]
                    else:
                        month_ranges[month][1] = col
    return month_ranges


def add_separator_column(ws, max_row, separator_column):
    """Add a black-filled separator column."""
    for row in range(1, max_row + 1):
        separator_cell = ws.cell(row=row, column=separator_column)
        separator_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.column_dimensions[ws.cell(row=1, column=separator_column).column_letter].width = 3


def apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row):
    """Apply formatting to the worksheet."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for col_num in range(2, max_col + 1):
        column_title = ws.cell(row=1, column=col_num).value
        if column_title:
            day_part = column_title.split(" ")[0]
            if day_part in DAY_COLORS:
                ws.cell(row=1, column=col_num).fill = PatternFill(
                    start_color=DAY_COLORS[day_part], end_color=DAY_COLORS[day_part], fill_type="solid"
                )

    for row in [total_rooms_row, total_camping_row]:
        if row:
            for col in range(1, max_col + 2):
                ws.cell(row=row, column=col).fill = YELLOW_FILL
                ws.cell(row=row, column=col).font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col + 1):
        for cell in row:
            cell.border = THIN_BORDER

    ws.freeze_panes = "B2"


def apply_excel_formatting_and_formulas(output_file, year):
    """Apply formatting and formulas to the output Excel file."""
    wb = load_workbook(output_file)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    total_rooms_row = total_camping_row = None

    for row in range(2, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == f"Total Rooms":
            total_rooms_row = row
        elif cell_value == f"Total Camping":
            total_camping_row = row

    apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col)
    apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row, year=year)
    apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row)

    # **Find and remove only columns with "Παρ 02/05" or "Fri 02/05" format**
    date_pattern = re.compile(r"^(Δευ|Τρι|Τετ|Πεμ|Παρ|Σαβ|Κυρ|Mon|Tue|Wed|Thu|Fri|Sat|Sun) \d{2}/\d{2}$")
    date_cols = [
        col for col in range(2, ws.max_column + 1)
        if isinstance(ws.cell(row=1, column=col).value, str) and date_pattern.match(ws.cell(row=1, column=col).value)
    ]

    # **Drop detected date columns LAST without affecting structure**
    for col in reversed(date_cols):  # Delete from right to left to avoid shifting issues
        ws.delete_cols(col)

    wb.save(output_file)


def per_nat_stage2(input_file, output_file, year):
    """Process reservations and generate the output Excel file."""
    logger.info(f'Starting with Stage 6. Year: {year}. Input File: {input_file}')
    df, headers = load_and_prepare_data(input_file)
    df = format_dates(df)
    split_index = find_camping_first_index(df)
    df = insert_totals_and_spacing(df, split_index, year=year)
    df.to_excel(output_file, index=False, engine='openpyxl')
    apply_excel_formatting_and_formulas(output_file, year=year)
    logger.info(f'Stage 6 completed. File saved as {output_file}')


if __name__ == "__main__":
    # Default file paths (for standalone execution)
    INPUT_FILE = "availabilityPerNationality2024.xls"
    OUTPUT_FILE = "per_nat_stage2_output_2024.xlsx"

    # Run stage6
    per_nat_stage2(INPUT_FILE, OUTPUT_FILE, 2024)
