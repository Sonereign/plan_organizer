import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from logger import logger

# Constants
HOUSE_KEYWORDS = ["Beach Apt", ".LUX for 4", ".Safari Tent 5pax", ".Sea Safari 4pax",
                  ".Skyline 3pax", ".Standard Mobile Home", ".ΤΡΟΧΟΣΠΙΤΑ DELUXE",
                  ".ΤΡΟΧΟΣΠΙΤΑ SEA VIEW", ".ΤΡΟΧΟΣΠΙΤΑ standard", ".Beach Apt / for 5.2.6", ".LUX Mobile Homes",
                  ".Mobile Home", ".Safari 5pax", ".Safari SST 4pax"]
YOUTH_HOSTEL_KEYWORDS = ["Youth Hostel"]

DAY_COLORS = {
    "Fri": "ADD8E6",  # Light Blue (Friday)
    "Sat": "90EE90",  # Light Green (Saturday)
    "Sun": "FFB6C1"  # Light Pink (Sunday)
}


def load_data(input_file):
    """Load Excel file and extract first sheet."""
    try:
        logger.info(f"Loading file: {input_file}")
        df = pd.read_excel(input_file, sheet_name=None)
        sheet_name = list(df.keys())[0]  # Get first sheet
        df = df[sheet_name]
        df.iloc[:, 0] = df.iloc[:, 0].astype(str)  # Ensure Category column is a string
        logger.debug("File loaded successfully.")
        return df
    except Exception as e:
        logger.error(f"Error loading file {input_file}: {e}")
        raise


def detect_date_columns(df, year):
    """Detect first and last date columns that match the given year."""
    try:
        logger.debug(f"Detecting date columns for year {year}.")

        # Extract columns that are valid dates and match the given year
        date_cols = [col for col in df.columns if is_date(col) and str(year) in str(col)]

        if date_cols:
            logger.debug(f"Date columns detected for {year}: {date_cols}")
            return date_cols[0], date_cols[-1]  # First and last date column
        else:
            logger.warning(f"No date columns found for year {year}.")
            return None, None
    except Exception as e:
        logger.error(f"Error detecting date columns: {e}")
        raise


def is_date(column_name):
    """Check if a column name is a date."""
    try:
        pd.to_datetime(column_name, dayfirst=True)
        return True
    except (ValueError, TypeError):
        return False


def format_date_columns(df, first_date_col, last_date_col, year):
    """Format date columns from first to last to 'Mon 14/9/YYYY' using the provided year."""
    try:
        if first_date_col and last_date_col:
            logger.info(f"Formatting date columns for year {year}.")
            date_range = df.loc[:, first_date_col:last_date_col].columns
            df.rename(columns={col: pd.to_datetime(col).strftime("%a %d/%m/%Y") for col in date_range}, inplace=True)
            logger.debug("Date columns formatted successfully.")
    except Exception as e:
        logger.error(f"Error formatting date columns: {e}")
        raise


def contains_keyword(value, keywords):
    """Check if a value contains any keyword from a list."""
    if pd.isna(value):
        return False
    value = value.strip()  # Remove leading/trailing spaces
    return any(keyword.lower() in value.lower() for keyword in keywords)


def add_empty_separator(df):
    """Create an empty row separator matching the number of columns in df."""
    return pd.DataFrame([[""] * df.shape[1]], columns=df.columns)


def split_sections(df):
    """Split the dataframe into Houses, Youth Hostel, and The Rest."""
    try:
        logger.info("Splitting data into sections.")
        houses = df[df.iloc[:, 0].apply(lambda x: contains_keyword(x, HOUSE_KEYWORDS))]
        youth_hostel = df[df.iloc[:, 0].apply(lambda x: contains_keyword(x, YOUTH_HOSTEL_KEYWORDS))]
        the_rest = df[~df.index.isin(houses.index) & ~df.index.isin(youth_hostel.index)]
        logger.debug("Sections split successfully.")
        return pd.concat([houses, add_empty_separator(df), youth_hostel, add_empty_separator(df), the_rest],
                         ignore_index=True)
    except Exception as e:
        logger.error(f"Error splitting sections: {e}")
        raise


def calculate_totals(section_df, category_col, capacity_col, group_name, year):
    """Calculate totals for a section and return the updated DataFrame with formatted total row."""
    if section_df.empty:
        return section_df

    # Select only numerical columns (skip category and capacity)
    numerical_data = section_df.iloc[:, 2:].apply(pd.to_numeric, errors='coerce')

    # Sum each date column
    totals_row = pd.DataFrame(numerical_data.sum()).T

    # Format the total row label
    total_label = f"Total {group_name} {year}"
    totals_row.insert(0, capacity_col, "")  # Insert 'TOTAL' in the second column (capacity)
    totals_row.insert(0, category_col, total_label)  # Insert formatted label in the first column (category)

    # Reset index for correct placement
    section_df = section_df.reset_index(drop=True)
    totals_row.index = [len(section_df)]  # Place at the end of section

    return pd.concat([section_df, totals_row], ignore_index=True)


def split_sections_with_totals(df, year):
    """Split the dataframe into sections, add a totals row for each section, and insert separators."""
    try:
        logger.info("Splitting data and adding totals.")

        category_col = df.columns[0]
        capacity_col = df.columns[1]

        # Ensure all category values are strings and strip spaces
        df[category_col] = df[category_col].astype(str).str.strip()

        # DEBUG: Print unique categories before filtering
        logger.debug(f"Unique categories before filtering: {df[category_col].unique().tolist()}")

        # Create boolean masks for filtering
        house_mask = df[category_col].str.contains('|'.join(map(re.escape, HOUSE_KEYWORDS)), case=False, na=False)
        hostel_mask = df[category_col].str.contains('|'.join(map(re.escape, YOUTH_HOSTEL_KEYWORDS)), case=False,
                                                    na=False)

        # Filter sections
        houses = df[house_mask]
        youth_hostel = df[hostel_mask]
        the_rest = df[~(house_mask | hostel_mask)]

        # Add totals to each section with formatted labels
        houses = calculate_totals(houses, category_col, capacity_col, "Accommodation", year)
        youth_hostel = calculate_totals(youth_hostel, category_col, capacity_col, "Youth Hostel", year)
        the_rest = calculate_totals(the_rest, category_col, capacity_col, "Camping", year)

        # Combine sections with empty separator rows
        result = pd.concat([
            houses, add_empty_separator(df),
            youth_hostel, add_empty_separator(df),
            the_rest
        ], ignore_index=True)

        logger.info("Sections split and totals added successfully.")
        return result

    except Exception as e:
        logger.error(f"Error processing sections with totals: {e}")
        raise


def save_to_excel(df, output_file):
    """Save dataframe to an Excel file."""
    try:
        logger.info(f"Saving data to {output_file}")
        df.to_excel(output_file, index=False, engine='openpyxl')
        logger.debug("Data saved successfully.")
    except Exception as e:
        logger.error(f"Error saving file {output_file}: {e}")
        raise


def apply_day_colors(output_file):
    """Apply colors only to the date header cells based on DAY_COLORS."""
    try:
        logger.info("Applying colors to date headers.")
        wb = load_workbook(output_file)
        ws = wb.active  # Get the active sheet
        headers = [cell.value for cell in ws[1]]
        for col_idx, col_name in enumerate(headers, start=1):
            if isinstance(col_name, str) and len(col_name) > 3:
                day = col_name[:3]  # Extract day part (e.g., "Mon")
                if day in DAY_COLORS:
                    fill = PatternFill(start_color=DAY_COLORS[day], end_color=DAY_COLORS[day], fill_type="solid")
                    ws.cell(row=1, column=col_idx).fill = fill
        wb.save(output_file)
        logger.debug("Coloring applied successfully.")
    except Exception as e:
        logger.error(f"Error applying colors: {e}")
        raise


def keep_only_totals(df):
    """Iterate from row index 1, keeping only rows where the first cell contains 'Total'."""
    try:
        logger.info("Filtering only total rows and header.")

        category_col = df.columns[0]  # Identify the first column (Category)
        index = 0  # Start from row index 1 (header at 0 is always kept)

        while index < len(df):
            first_cell = str(df.iloc[index, 0])  # Read first cell in the row

            logger.debug(f"Checking row {index}: '{first_cell}'")

            if first_cell.startswith("Total") or first_cell.startswith("Category"):
                logger.debug(f"✅ Keeping row {index}: '{first_cell}'")
                index += 1  # Move to next row
            else:
                logger.debug(f"❌ Dropping row {index}: '{first_cell}'")
                df.drop(index, inplace=True)
                df.reset_index(drop=True, inplace=True)  # Reset index after drop

        logger.debug(f"Final categories in dataset: {df[category_col].tolist()}")
        logger.info("Filtered dataset to keep only total rows and header.")

        return df
    except Exception as e:
        logger.error(f"Error filtering total rows and header: {e}")
        raise


def per_zone_per_type_stage5_previous_years(input_file, output_file, year):
    logger.debug(f'Processing {input_file}')

    try:
        df = load_data(input_file)

        first_date_col, last_date_col = detect_date_columns(df, year)
        logger.info(f"Date range detected: {first_date_col} - {last_date_col}")

        if first_date_col and last_date_col:
            format_date_columns(df, first_date_col, last_date_col, year)

        # Process sections with totals
        df_split = split_sections_with_totals(df, year)

        # Keep only header and totals
        df_totals_only = keep_only_totals(df_split)

        save_to_excel(df_totals_only, output_file)
        apply_day_colors(output_file)

        logger.info(f"Stage 5 perZone {year} completed. File saved as {output_file}")
    except Exception as e:
        logger.error(f"Stage 5 perZone {year}: {e}")
        raise


if __name__ == "__main__":
    INPUT_FILE = "sources/availabilityPerZone2023.xls"
    OUTPUT_FILE = "per_zone_stage5_output_2023.xlsx"

    # Run stage4
    per_zone_per_type_stage5_previous_years(INPUT_FILE, OUTPUT_FILE, 2023)