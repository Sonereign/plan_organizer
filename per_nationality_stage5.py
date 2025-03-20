import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from logger import logger


def get_headers(ws):
    """Returns a list of headers from the first row of the worksheet."""
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    logger.debug(f"Extracted headers: {headers}")
    return headers


def find_last_category_column(headers):
    """Finds the last occurrence of the 'Category' column."""
    for col in range(len(headers), 0, -1):
        if headers[col - 1] == "Category":
            logger.debug(f"Last 'Category' column found at index {col}")
            return col
    logger.warning("'Category' column not found.")
    return None


def delete_columns_after_category(ws, last_category_index):
    """Deletes all columns after the last occurrence of 'Category'."""
    try:
        if last_category_index:
            ws.delete_cols(last_category_index + 1, ws.max_column - last_category_index)
            logger.info(f"Deleted columns from index {last_category_index + 1} onward.")
        else:
            logger.warning("No columns deleted as 'Category' column was not found.")
    except Exception as e:
        logger.exception("Error deleting columns after 'Category': %s", e)


def add_percentage_columns(ws, previous_years):
    """Adds 'Percent to Total YYYY' columns for the current year and given previous years."""
    try:
        headers = get_headers(ws)
        last_col = len(headers) + 1  # Insert after the last existing column
        current_year = datetime.now().year
        all_years = [current_year] + previous_years  # Include current year first

        for index, year in enumerate(all_years):
            ws.insert_cols(last_col + index)
            ws.cell(row=1, column=last_col + index).value = f"Percent to Total {year}"

        logger.info(f"Added 'Percent to Total' columns for: {all_years}")
        return last_col + len(all_years)
    except Exception as e:
        logger.exception("Error adding percentage columns: %s", e)


def find_and_replace_percent_to_total_column(ws, current_year):
    """
    Finds the first occurrence of the column 'Percent to Total {current_year}',
    replaces it with a black-filled separator, and returns the next available column index.
    """
    try:
        headers = get_headers(ws)
        target_header = f"Percent to Total {current_year}"

        for col in range(1, len(headers) + 1):
            if ws.cell(row=1, column=col).value == target_header:
                ws.cell(row=1, column=col).value = ""
                black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = black_fill
                logger.info(f"Replaced '{target_header}' column at position {col} with a black-filled separator.")
                return col + 1

        logger.info(f"Column '{target_header}' not found.")
        return None
    except Exception as e:
        logger.exception("Error replacing 'Percent to Total' column: %s", e)


def add_separator_column(ws, insert_at_col):
    """Adds a black-filled separator column at the specified position."""
    try:
        ws.insert_cols(insert_at_col)
        ws.cell(row=1, column=insert_at_col).value = ""
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=insert_at_col).fill = black_fill

        logger.info(f"Added a black-filled separator column at position {insert_at_col}")
        return insert_at_col + 1
    except Exception as e:
        logger.exception("Error adding separator column: %s", e)


def add_percent_difference_columns(ws, insert_at_col, previous_years):
    """Adds 'Percent difference current_year - previous_year' columns after the first separator."""
    try:
        current_year = datetime.now().year
        for index, prev_year in enumerate(previous_years):
            ws.insert_cols(insert_at_col + index)
            ws.cell(row=1, column=insert_at_col + index).value = f"Percent difference {current_year} - {prev_year}"

        logger.info(f"Added 'Percent difference' columns for: {previous_years}")
        return insert_at_col + len(previous_years)
    except Exception as e:
        logger.exception("Error adding percent difference columns: %s", e)


def process_stage9(input_file, output_file, previous_years):
    """Processes Stage 9 by deleting columns after 'Category', adding percentage columns, separators, and percent differences."""
    try:
        logger.info("#######################################################")
        logger.info(
            f"Running Stage 9 with input_file={input_file}, output_file={output_file}, previous_years={previous_years}")
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        headers = get_headers(ws)

        next_available_col = add_percentage_columns(ws, previous_years)
        next_available_col = add_separator_column(ws, next_available_col)
        next_available_col = add_percent_difference_columns(ws, next_available_col, previous_years)
        add_separator_column(ws, next_available_col)

        current_year = datetime.now().year
        find_and_replace_percent_to_total_column(ws, current_year)

        wb.save(output_file)
        logger.info(f"Stage 9 processing complete. Output saved to {output_file}")
    except Exception as e:
        logger.exception("Error processing Stage 9: %s", e)


def per_nationality_stage5(input_path, output_path, previous_years):
    """Entry point for Stage 9 processing."""
    try:
        process_stage9(input_file=input_path, output_file=output_path, previous_years=previous_years)
    except Exception as e:
        logger.exception("Error in per_nationality_stage5: %s", e)


if __name__ == '__main__':
    input_path = "nat_stage4_output.xlsx"
    output_path = "nat_stage5_output.xlsx"
    previous_years = [2024, 2023]  # Example: List of previous years
    per_nationality_stage5(input_path=input_path, output_path=output_path, previous_years=previous_years)
