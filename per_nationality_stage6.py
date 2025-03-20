import re
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter  # Convert column index to Excel letters
from logger import logger

# Define the months for reference
MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Define six distinct blue shades for different years
BLUE_SHADES = ["538DD5", "8DB4E2", "C5D9F1", "4F81BD", "95B3D7", "DCE6F1"]


def find_total_current_year_column(ws, total_column):
    """Find the column index of 'Total current year'."""
    current_year = datetime.now().year

    try:
        for col in range(2, total_column + 1):
            header_value = ws.cell(row=1, column=col).value
            col_letter = get_column_letter(col)

            if isinstance(header_value, str) and header_value.strip() == f"Total {current_year}":
                logger.info(f"Found 'Total {current_year}' column at {col_letter} (Excel Col {col})")
                return col

        logger.warning(f"'Total {current_year}' column not found!")
    except Exception as e:
        logger.exception("Error occurred while finding 'Total current year' column.")

    return None


def find_monthly_column_ranges(ws, total_column):
    """Find the first and last column for each month based on daily entries."""
    month_ranges = {}

    try:
        for col in range(2, total_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if not isinstance(header_value, str):
                continue

            for i, month in enumerate(MONTHS):
                month_number = str(i + 4).zfill(2)  # "Apr" = 04, "May" = 05, etc.
                if header_value.endswith(f'/{month_number}'):
                    if month not in month_ranges:
                        month_ranges[month] = [col, col]  # Start and end at the same column
                    else:
                        month_ranges[month][1] = col  # Update end column

        logger.debug(f"Monthly column ranges identified: {month_ranges}")
    except Exception as e:
        logger.exception(f"Error occurred while finding monthly column ranges. {e}")

    return month_ranges


def find_existing_monthly_columns(ws, total_column):
    """Find columns labeled 'Apr current_year', 'May current_year', etc."""
    monthly_columns = {}
    current_year = datetime.now().year

    try:
        for col in range(2, total_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if isinstance(header_value, str) and header_value in [f"{month} {current_year}" for month in MONTHS]:
                monthly_columns[header_value] = col
                logger.info(f"Found '{header_value}' at column {get_column_letter(col)} (Excel Col {col})")

        logger.debug(f"Existing monthly columns: {monthly_columns}")
    except Exception as e:
        logger.exception("Error occurred while finding existing monthly columns.")

    return monthly_columns


def should_skip_row(ws, row):
    """Check if the row should be skipped based on its first cell value."""
    first_cell_value = ws.cell(row=row, column=1).value
    return isinstance(first_cell_value, str) and first_cell_value in ["pan_pan", "sep_row"]


def insert_monthly_sums(ws, max_row, month_ranges, monthly_columns, total_rooms_row, total_camping_row):
    """Insert sum formulas into existing 'Apr current_year', 'May current_year', etc. columns."""
    logger.info("Placing sum formulas in existing monthly columns...")
    current_year = datetime.now().year

    try:
        for month, (start_col, end_col) in month_ranges.items():
            sum_col = monthly_columns.get(f"{month} {current_year}")
            if not sum_col:
                logger.warning(f"No column found for {month} {current_year}. Skipping.")
                continue

            sum_col_letter = get_column_letter(sum_col)
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)

            logger.info(f"Adding sums for {month}: Daily Columns {start_letter} to {end_letter}, Sum in {sum_col_letter}")

            for row in range(2, max_row + 1):
                if row not in [total_rooms_row, total_camping_row] and not should_skip_row(ws, row):
                    sum_formula = f"=SUM({start_letter}{row}:{end_letter}{row})"
                    ws.cell(row=row, column=sum_col).value = sum_formula
                    logger.debug(f"Row {row} (Excel {sum_col_letter}{row}): {sum_formula}")
    except Exception as e:
        logger.exception(f"Error occurred while inserting monthly sum formulas. {e}")

def insert_total_sums(ws, max_row, monthly_columns, total_current_year_col, total_rooms_row, total_camping_row):
    """Insert sum formulas into 'Total current_year' column, summing all monthly columns and coloring the cell yellow."""
    current_year = datetime.now().year

    try:
        if not total_current_year_col:
            logger.warning(f"'Total {current_year}' column not found. Skipping.")
            return

        total_current_year_letter = get_column_letter(total_current_year_col)
        sum_range_letters = [get_column_letter(col) for col in monthly_columns.values()]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)

        logger.info(f"Adding total sums in 'Total {current_year}' column ({total_current_year_letter})...")

        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row] and not should_skip_row(ws, row):
                sum_formula = f"=SUM({','.join([f'{col}{row}' for col in sum_range_letters])})"
                cell = ws.cell(row=row, column=total_current_year_col)
                cell.value = sum_formula
                cell.fill = yellow_fill
                cell.font = bold_font
                logger.debug(f"Row {row} (Excel {total_current_year_letter}{row}): {sum_formula}")
    except Exception as e:
        logger.exception(f"Error occurred while inserting total sum formulas. {e}")


def add_monthly_sums(ws, max_row, total_column, total_rooms_row, total_camping_row):
    """Find the necessary columns and insert sum formulas."""
    try:
        month_ranges = find_monthly_column_ranges(ws, total_column)
        monthly_columns = find_existing_monthly_columns(ws, total_column)
        total_current_year_col = find_total_current_year_column(ws, total_column)

        insert_monthly_sums(ws, max_row, month_ranges, monthly_columns, total_rooms_row, total_camping_row)
        insert_total_sums(ws, max_row, monthly_columns, total_current_year_col, total_rooms_row, total_camping_row)
    except Exception as e:
        logger.exception(f"Error occurred in add_monthly_sums function. {e}")


def find_total_rows(ws):
    """Find the row indices for 'Total Rooms' and 'Total Camping'."""
    total_rooms_row = total_camping_row = None
    try:
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "Total Rooms":
                total_rooms_row = row
            elif cell_value == "Total Camping":
                total_camping_row = row
        logger.debug(f"Total Rooms Row: {total_rooms_row}, Total Camping Row: {total_camping_row}")
    except Exception as e:
        logger.error(f"Error finding total rows: {e}", exc_info=True)
    return total_rooms_row, total_camping_row


def should_stop_summing(ws, row):
    """Check if the summing should stop based on the first cell value."""
    try:
        first_cell_value = ws.cell(row=row, column=1).value
        return isinstance(first_cell_value, str) and first_cell_value in ["pan_pan", "sep_row"]
    except Exception as e:
        logger.error(f"Error in should_stop_summing at row {row}: {e}", exc_info=True)
        return False


def is_black_filled(ws, row, col):
    """Check if a cell or its adjacent cells (above or below) are filled with black color."""
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    def check_fill(r, c):
        fill = ws.cell(row=r, column=c).fill
        return isinstance(fill, PatternFill) and fill.start_color.rgb == "00000000"

    try:
        if check_fill(row, col):
            logger.debug(f"{get_column_letter(col)}{row} is already black.")
            return True
        if row > 1 and check_fill(row - 1, col):
            ws.cell(row=row, column=col).fill = black_fill
            logger.debug(f"{get_column_letter(col)}{row} filled black due to upper cell.")
            return True
        if row < ws.max_row and check_fill(row + 1, col):
            ws.cell(row=row, column=col).fill = black_fill
            logger.debug(f"{get_column_letter(col)}{row} filled black due to lower cell.")
            return True
    except Exception as e:
        logger.error(f"Error checking black fill at {get_column_letter(col)}{row}: {e}", exc_info=True)
    return False


def insert_total_room_camping_sums(ws, total_column, total_row):
    """Insert sum formulas into 'Total Rooms' and 'Total Camping' rows."""
    if not total_row:
        logger.warning("Total row is None, skipping sum insertion.")
        return

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    try:
        for col in range(2, total_column + 1):
            col_letter = get_column_letter(col)

            if ws.cell(row=1, column=col).value is None:
                logger.debug(f"Skipping column {col_letter} as first row is empty.")
                continue

            sum_formula = "=SUM("
            start_row = total_row - 1

            while start_row > 1:
                if should_stop_summing(ws, start_row):
                    break
                start_row -= 1

            sum_range = []
            for r in range(start_row + 1, total_row):
                if not is_black_filled(ws, r, col):
                    sum_range.append(f"{col_letter}{r}")
                else:
                    logger.debug(f"Skipping {col_letter}{r} due to black fill.")

            if sum_range:
                sum_formula += ",".join(sum_range) + ")"
                cell = ws.cell(row=total_row, column=col)
                cell.value = sum_formula
                cell.fill = yellow_fill
                cell.font = bold_font
                logger.info(f"Inserted sum formula at {col_letter}{total_row}: {sum_formula}")
            else:
                logger.warning(f"No valid cells to sum for {col_letter}{total_row}")
    except Exception as e:
        logger.error(f"Error inserting total room/camping sums: {e}", exc_info=True)


def apply_grid_borders(ws):
    """Apply a full grid border to all data cells in the worksheet."""
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border


def find_column_by_header(ws, header_name):
    """Find the column index for a given header."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return col
    return None


def calculate_percent_to_total(ws, previous_years):
    """Calculate Percent to Total values."""
    try:
        current_year = datetime.now().year
        all_years = [current_year] + previous_years

        for year in all_years:
            percent_col = find_column_by_header(ws, f"Percent to Total {year}")
            total_col = find_column_by_header(ws, f"Total {year}")
            total_rooms_row, total_camping_row = find_total_rows(ws)

            if not percent_col or not total_col or (not total_rooms_row and not total_camping_row):
                logger.error("Required columns or rows not found for calculations!")
                return

            for row in range(2, ws.max_row + 1):
                total_cell = ws.cell(row=row, column=total_col)
                percent_cell = ws.cell(row=row, column=percent_col)
                category = ws.cell(row=row, column=1).value
                total_ref_row = total_rooms_row if "Rooms" in category else total_camping_row

                if total_ref_row:
                    total_ref_cell = ws.cell(row=total_ref_row, column=total_col)
                    if total_cell.value is not None and total_ref_cell.value is not None:
                        percent_cell.value = f"={total_cell.coordinate}/{total_ref_cell.coordinate}"
                        percent_cell.number_format = "0.00%"
                        logger.debug(f"Calculated {percent_cell.coordinate}: {percent_cell.value}")
    except Exception as e:
        logger.error(f"Error in calculate_percent_to_total: {e}", exc_info=True)


def calculate_percent_difference(ws, previous_years):
    """Calculate Percent Difference current_year - previous_years and apply conditional formatting."""
    try:
        current_year = datetime.now().year
        for year in previous_years:
            percent_diff_col = find_column_by_header(ws, f"Percent difference {current_year} - {year}")
            total_current_year_col = find_column_by_header(ws, f"Total {current_year}")
            total_previous_year_col = find_column_by_header(ws, f"Total {year}")

            if not percent_diff_col or not total_current_year_col or not total_previous_year_col:
                logger.error("[ERROR] Required columns not found!")
                return

            for row in range(2, ws.max_row + 1):
                total_current_year_cell = ws.cell(row=row, column=total_current_year_col)
                total_previous_year_cell = ws.cell(row=row, column=total_previous_year_col)
                percent_diff_cell = ws.cell(row=row, column=percent_diff_col)

                if total_current_year_cell.value is not None and total_previous_year_cell.value is not None:
                    percent_diff_cell.value = (
                        f"=IF({total_previous_year_cell.coordinate}<>0,"
                        f" ({total_current_year_cell.coordinate}-{total_previous_year_cell.coordinate})"
                        f"/{total_previous_year_cell.coordinate}, 0)"
                    )
                    percent_diff_cell.number_format = "0.00%"
                    logger.debug(f"{percent_diff_cell.coordinate} = {percent_diff_cell.value}")

            percent_diff_range = (
                f"{get_column_letter(percent_diff_col)}2:" 
                f"{get_column_letter(percent_diff_col)}{ws.max_row}"
            )
            color_scale_rule = ColorScaleRule(
                start_type="num", start_value=-1, start_color="FFCCCC",  # Red for negative
                mid_type="num", mid_value=0, mid_color="FFFFFF",  # White for neutral
                end_type="num", end_value=1, end_color="CCFFCC"  # Green for positive
            )
            ws.conditional_formatting.add(percent_diff_range, color_scale_rule)
            logger.info(f"Applied conditional formatting for {year}.")
    except Exception as e:
        logger.exception(f"Exception occurred in calculate_percent_difference: {e}")


def fill_black_columns(ws):
    """Find empty headers or 'sep_col' headers and apply black fill to the entire column."""
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col)
        if header_cell.value is None or str(header_cell.value).strip().lower() == "sep_col":
            logger.info(f"[DEBUG] Header '{header_cell.value}' in column {col} is blacked out.")
            for row in range(2, ws.max_row + 1):  # Fill all data rows
                ws.cell(row=row, column=col).fill = black_fill
                ws.column_dimensions[get_column_letter(col)].width = 5  # Set width to 5


def fill_black_rows(ws):
    """Find rows named 'pan_pan' or 'sep_row', and the row after the last data row, and black them out."""
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    last_data_row = ws.max_row

    for row in range(2, last_data_row + 1):
        first_cell = ws.cell(row=row, column=1)
        if first_cell.value is None or str(first_cell.value).strip().lower() in ["pan_pan", "sep_row"]:
            logger.info(f"[DEBUG] Row {row} ('{first_cell.value}') is blacked out.")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = black_fill

    # Black out the row after the last data row
    extra_row = last_data_row + 1
    logger.info(f"[DEBUG] Blacking out extra row {extra_row}.")
    for col in range(1, ws.max_column + 1):
        ws.cell(row=extra_row, column=col).fill = black_fill


def fill_date_columns(ws):
    """Find columns with 'Month YYYY' format and apply alternating blue shades for each year, skipping specific rows."""
    year_colors = {}  # Store assigned colors per year
    color_index = 0  # Track which color to assign next

    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col)
        header_value = header_cell.value

        if isinstance(header_value, str):
            match = re.match(r"([A-Za-z]{3}) (\d{4})", header_value)  # Match "Apr 2024" format
            if match:
                month, year = match.groups()
                if month in MONTHS:
                    # Assign color per year if not already assigned
                    if year not in year_colors:
                        year_colors[year] = BLUE_SHADES[color_index % len(BLUE_SHADES)]
                        color_index += 1  # Move to next color for next year

                    fill_color = PatternFill(start_color=year_colors[year], end_color=year_colors[year],
                                             fill_type="solid")

                    logger.info(f"[DEBUG] Coloring column {col} ({header_value}) with {year_colors[year]}")  # Debugging info

                    # Apply color to entire column, skipping specific rows
                    for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                        first_cell_value = ws.cell(row=row, column=1).value

                        # Skip rows where the first cell is empty, "pan_pan", "sep_row", or contains "Total"
                        if first_cell_value is None or first_cell_value in ["pan_pan", "sep_row"] or "Total" in str(
                                first_cell_value):
                            continue

                        ws.cell(row=row, column=col).fill = fill_color


def process_stage10(input_file, output_file, previous_years):
    """Processes Stage 10 by adding sum formulas to the input Excel file."""
    try:
        logger.info(f"Processing file: {input_file}")
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        max_row = ws.max_row
        total_column = ws.max_column
        total_rooms_row = total_camping_row = None

        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "Total Rooms":
                total_rooms_row = row
            elif cell_value == "Total Camping":
                total_camping_row = row

        logger.info(f"Max row: {max_row}")
        logger.info(f"Total Rooms Row: {total_rooms_row}")
        logger.info(f"Total Camping Row: {total_camping_row}")

        add_monthly_sums(ws, max_row, total_column, total_rooms_row, total_camping_row)
        insert_total_room_camping_sums(ws, total_column, total_rooms_row)
        insert_total_room_camping_sums(ws, total_column, total_camping_row)
        apply_grid_borders(ws)

        calculate_percent_to_total(ws, previous_years)
        calculate_percent_difference(ws, previous_years)
        fill_black_columns(ws)
        fill_black_rows(ws)
        fill_date_columns(ws)

        wb.save(output_file)
        logger.info(f"Stage 10 processing complete. Output saved to {output_file}")
    except Exception as e:
        logger.exception(f"Exception occurred in process_stage10: {e}")


def per_nationality_stage6(input_path, output_path, previous_years):
    """Entry point for Stage 10 processing."""
    try:
        process_stage10(input_file=input_path, output_file=output_path, previous_years=previous_years)
    except Exception as e:
        logger.exception(f"Exception occurred in per_nationality_stage6: {e}")


if __name__ == '__main__':
    input_path = "nat_stage5_output.xlsx"
    output_path = "nat_stage6_output.xlsx"
    previous_years = [2024, 2023]  # Example: List of previous years
    per_nationality_stage6(input_path=input_path, output_path=output_path, previous_years=previous_years)
