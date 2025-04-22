import os
import traceback
from copy import copy
from datetime import datetime
from tkinter import messagebox

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

import per_zone_stage6
from per_nat_stage1_finalizer import per_nat_stage1_finalizer
from per_zone_stage1 import per_zone_stage1
from per_zone_stage2 import per_zone_stage2
from per_zone_stage3 import per_zone_stage3
from per_zone_stage4 import per_zone_stage4
from per_zone_stage4_finalizer import per_zone_stage4_finalizer
from per_zone_stage5 import per_zone_per_type_stage5_previous_years
from per_zone_stage6 import per_zone_stage6
from per_nat_stage1 import per_nat_stage1
from per_nat_stage2 import per_nat_stage2
from per_nat_stage3 import per_nat_stage3
from per_nat_stage4 import per_nat_stage4
from per_nat_stage5 import per_nat_stage5
from per_nat_stage6 import per_nat_stage6
from logger import logger
from per_zone_stage7 import per_zone_stage7


def process_files(app):
    try:
        # Generate final output file name and sheet names
        today = datetime.today().strftime("%d-%m-%y")
        final_output = f"{today}_availabilityPerZone&Nationality.xlsx"
        sheet1_name = f"{today}-πληρότητα-units"
        sheet2_name = "εθνικότητες"

        full_zone = False
        no_zone = False

        per_zone_stage1_output = "per_zone_stage1_output.xlsx"
        per_zone_stage2_output = "per_zone_stage2_output.xlsx"
        per_zone_stage3_output = "per_zone_stage3_output.xlsx"
        per_zone_stage4_output = "per_zone_stage4_output.xlsx"
        per_zone_stage4_finalizer_output = "per_zone_stage4_finalizer_output.xlsx"
        per_zone_stage5_output_filenames = []
        per_zone_stage6_output = "per_zone_stage6_output.xlsx"
        per_zone_stage7_output = "per_zone_stage7_output.xlsx"

        per_nat_stage1_output = "per_nat_stage1_output.xlsx"
        per_nat_stage1_finalizer_output = "per_nat_stage1_finalizer_output.xlsx"
        per_nat_stage2_output_filenames = []
        per_nat_stage3_output = "per_nat_stage3_output.xlsx"
        per_nat_stage4_output = "per_nat_stage4_output.xlsx"
        per_nat_stage5_output = "per_nat_stage5_output.xlsx"
        per_nat_stage6_output = "per_nat_stage6_output.xlsx"

        if app.availability_per_type_path is None and app.availability_per_zone_path is None and app.availability_per_nationality_path is None:
            app.status_label.config(
                text="You know, sometimes you need to put some effort as well.. Please give me the paths to the files.")
            messagebox.showerror("ER0R!1!1 S0S",
                                 f"Αγαπητέ Λεωνίδα, θα κάνω οτι δεν είδα οτι ξέχασες να επιλέξεις αρχεία..")
            return
        if app.availability_per_zone_path is None or app.availability_per_type_path is None:
            app.status_label.config(
                text="Availability Per Zone will not be processed on this session because the \npath for Availability per Zone or Availability per Type is empty.")
            messagebox.showwarning("Warning",
                                   f"Availability Per Zone will not be processed on this session because the path for Availability per Zone or Availability per Type is empty.")
            no_zone = True
        else:
            per_zone_stage1(app.availability_per_zone_path, per_zone_stage1_output)
            per_zone_stage2(per_zone_stage1_output, app.availability_per_type_path, per_zone_stage2_output)
            per_zone_stage3(per_zone_stage2_output, per_zone_stage3_output)
            per_zone_stage4(per_zone_stage3_output, per_zone_stage4_output)

            # Run per_zone_stage5 for previous years
            for year, file_path in app.previous_years_zone_paths.items():
                output_file = f"per_zone_stage5_output_{year}.xlsx"
                per_zone_stage5_output_filenames.append(output_file)  # Append file name to list
                per_zone_per_type_stage5_previous_years(input_file=file_path, output_file=output_file, year=year)

            if not per_zone_stage5_output_filenames:
                """Calculate results for per_zone_stage4_finalizer_output without previous years"""
                per_zone_stage4_finalizer(per_zone_stage3_output, per_zone_stage4_finalizer_output)
            else:
                """Process previous years zone files"""
                per_zone_stage6(per_zone_stage4_output, per_zone_stage5_output_filenames, per_zone_stage6_output)
                per_zone_stage7(per_zone_stage6_output, per_zone_stage7_output)
                full_zone = True

        # Run per_nat_stage1 if nationality file is provided
        if app.availability_per_nationality_path:
            # Run per_nat_stage2 for previous years
            for year, file_path in app.previous_years_nat_paths.items():
                output_file = f"per_nat_stage2_output_{year}.xlsx"
                per_nat_stage2_output_filenames.append(output_file)  # Append file name to list
                per_nat_stage2(input_file=file_path, output_file=output_file, year=year)

            if not per_nat_stage2_output_filenames:
                per_nat_stage1_finalizer(app.availability_per_nationality_path, per_nat_stage1_finalizer_output)

                if no_zone:
                    """No zone data will be computed, only availabilityPerNationality"""
                    app.status_label.config(
                        text="COME FROM THIS SIDE SIIIIIIIIIIIIIIIR!!!.")
                    messagebox.showwarning("Warning",
                                           f"The developer was too lazy to allow you process only perNationality, you're getting nothing.\nUncheck Enable Cleanup and open per_nat_stage1_finalizer_output.xlsx")
                else:
                    if full_zone:
                        """Combine per_nat_stage1_finalizer_output.xlsx with per_zone_stage7_output.xlsx"""
                        combine_sheets(per_zone_final_file=per_zone_stage7_output,
                                       per_nat_final_file=per_nat_stage1_finalizer_output,
                                       final_output_name=final_output,
                                       sheet1_name=sheet1_name, sheet2_name=sheet2_name, app=app)
                        app.status_label.config(
                            text="Processing complete! Plan has per_zone prev year data and current year nationality data.")
                        messagebox.showinfo("Success",
                                            f"Plan has per_zone prev year data and current year per_nat data.\nFinal output saved as {final_output}")
                    else:
                        """Combine per_nat_stage1_finalizer_output.xlsx with per_zone_stage4_finalizer_output.xlsx"""
                        combine_sheets(per_zone_final_file=per_zone_stage4_finalizer_output,
                                       per_nat_final_file=per_nat_stage1_finalizer_output,
                                       final_output_name=final_output,
                                       sheet1_name=sheet1_name, sheet2_name=sheet2_name, app=app)
                        app.status_label.config(
                            text="Processing complete! Plan has data only for current year per_zone and per_nat.")
                        messagebox.showinfo("Success",
                                            f"Plan has data only for current year per_zone and per_nat.\nFinal output saved as {final_output}")
            else:
                # Create final output by combining sheets from stage4 and stage10 outputs
                per_nat_stage1(app.availability_per_nationality_path, per_nat_stage1_output)

                nat_previous_years = list(app.previous_years_nat_paths.keys())  # Extract years from dictionary keys
                nat_number_of_previous_year_data = len(per_nat_stage2_output_filenames)

                per_nat_stage3(per_nat_stage1_output, per_nat_stage2_output_filenames, per_nat_stage3_output)
                per_nat_stage4(per_nat_stage3_output, per_nat_stage4_output, nat_previous_years,
                               nat_number_of_previous_year_data)
                per_nat_stage5(per_nat_stage4_output, per_nat_stage5_output, nat_previous_years)
                per_nat_stage6(per_nat_stage5_output, per_nat_stage6_output, nat_previous_years)

                if full_zone:
                    """We need to merge per_zone_stage7 and per_nat_stage6"""
                    combine_sheets(per_zone_stage7_output, per_nat_stage6_output, final_output, sheet1_name,
                                   sheet2_name, app=app)
                    app.status_label.config(
                        text="Processing complete! Plan has prev_year_data for both per_zone and per_nat.\n")
                    messagebox.showinfo("Success",
                                        f"Plan has prev_year_data for both per_zone and per_nat.\nFinal output saved as {final_output}")
                else:
                    """We need to make calculations for per_zone_stage4_finalizer_output and then combine with per_nat_stage6"""
                    combine_sheets(per_zone_stage4_finalizer_output, per_nat_stage6_output, final_output, sheet1_name,
                                   sheet2_name, app=app)
                    app.status_label.config(
                        text="Processing complete! Plan has prev year data for per_nat but current year data for per_zone.\n")
                    messagebox.showinfo("Success",
                                        f"Plan has prev year data for per_nat but current year data for per_zone.\nFinal output saved as {final_output}")
        else:
            logger.info(f'No path given for Nationality current year, no need to combine, just pack zones')
            if full_zone:
                """We need just to rename the stage7 output to date_availabilityPerZone.xlsx"""
                combine_sheets(per_zone_final_file=per_zone_stage7_output,
                               per_nat_final_file=None,
                               final_output_name=f"{today}_availabilityPerZone&PreviousYears.xlsx",
                               sheet1_name=sheet1_name, sheet2_name=None, app=app)
                app.status_label.config(
                    text="Processing complete! Plan has only per_zone and prev years data.\n")
                messagebox.showinfo("Success",
                                    f"Plan has only per_zone and prev years data.\nFinal output saved as {final_output}")
            else:
                """We need to make calculations on per_zone_stage4 and have an output date_availabilityPerZone.xlsx"""
                combine_sheets(per_zone_final_file=per_zone_stage4_finalizer_output,
                               per_nat_final_file=None,
                               final_output_name=f"{today}_availabilityPerZone.xlsx",
                               sheet1_name=sheet1_name, sheet2_name=None, app=app)
                app.status_label.config(
                    text="Processing complete! Plan has only per_zone current year data.\n")
                messagebox.showinfo("Success",
                                    f"Plan has only per_zone current year data.\nFinal output saved as {final_output}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        logger.error(f"An error occurred: {e} {traceback.format_exc()}")
    finally:
        app.process_button.config(state="normal")
        if app.cleanup_var:
            # Clean up temporary files
            for file in [per_zone_stage1_output,
                         per_zone_stage2_output,
                         per_zone_stage3_output,
                         per_zone_stage4_output,
                         per_zone_stage4_finalizer_output,
                         per_zone_stage6_output,
                         per_zone_stage7_output,
                         per_nat_stage1_output,
                         per_nat_stage1_finalizer_output,
                         per_nat_stage3_output,
                         per_nat_stage4_output,
                         per_nat_stage5_output,
                         per_nat_stage6_output ]:
                if os.path.exists(file):
                    os.remove(file)
            for file in per_nat_stage2_output_filenames:
                if os.path.exists(file):
                    os.remove(file)
            for file in per_zone_stage5_output_filenames:
                if os.path.exists(file):
                    os.remove(file)

def combine_sheets(per_zone_final_file, per_nat_final_file, final_output_name, sheet1_name, sheet2_name, app):
    """Combine sheets from stage4 and stage5 outputs into a single Excel file."""
    # Load workbooks
    wb_stage5 = None

    if per_nat_final_file is not None:
        wb_stage5 = load_workbook(per_nat_final_file) if app.availability_per_nationality_path else None

    # Create a new workbook for the final output
    wb_final = load_workbook(per_zone_final_file)

    # Rename the sheet from stage4 to the custom sheet1 name
    sheet_stage4 = wb_final.active
    sheet_stage4.title = sheet1_name

    # Add sheet from stage5 if available
    if wb_stage5 is not None:
        sheet_stage5 = wb_stage5.active
        # Create a new sheet in the final workbook for Stage5 Results
        new_sheet_stage5 = wb_final.create_sheet(sheet2_name)

        # Copy all cells, styles, and formulas from Stage5 sheet to the new sheet
        for row in sheet_stage5.iter_rows():
            for cell in row:
                new_cell = new_sheet_stage5.cell(
                    row=cell.row, column=cell.column, value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy(cell.font)  # Use copy function
                    new_cell.border = copy(cell.border)  # Use copy function
                    new_cell.fill = copy(cell.fill)  # Use copy function
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)  # Use copy function
                    new_cell.alignment = copy(cell.alignment)  # Use copy function

    # Save the final workbook
    wb_final.save(final_output_name)
    if wb_stage5 is not None:
        apply_conditional_formatting(final_output_name)


def apply_conditional_formatting(file_path):
    # Load the workbook and select the specified worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb["εθνικότητες"]

    ws.freeze_panes = "B2"

    # Find columns that start with "Percent difference"
    header_row = ws[1]  # Assuming headers are in the first row
    percent_diff_cols = [cell.column for cell in header_row if
                         cell.value and cell.value.startswith("Percent difference")]

    # Apply conditional formatting to each identified column
    for percent_diff_col in percent_diff_cols:
        percent_diff_range = f"{get_column_letter(percent_diff_col)}2:{get_column_letter(percent_diff_col)}{ws.max_row}"
        color_scale_rule = ColorScaleRule(
            start_type="num", start_value=-1, start_color="FFCCCC",  # Red for negative
            mid_type="num", mid_value=0, mid_color="FFFFFF",  # White for neutral
            end_type="num", end_value=1, end_color="CCFFCC"  # Green for positive
        )
        ws.conditional_formatting.add(percent_diff_range, color_scale_rule)

    # Save the workbook
    wb.save(file_path)
    logger.info("Conditional formatting applied successfully!")
