import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment
from logger import logger


def load_stage5_data(stage5_file):
    wb5 = openpyxl.load_workbook(stage5_file)
    ws5 = wb5.active
    countries_stage5 = {ws5.cell(row=row, column=1).value: row for row in range(2, ws5.max_row + 1) if
                        ws5.cell(row=row, column=1).value}
    empty_row = next((row for row in range(2, ws5.max_row + 1) if ws5.cell(row=row, column=1).value is None), None)

    if empty_row:
        for col in range(1, ws5.max_column + 1):
            ws5.cell(row=empty_row, column=col, value="sep_row")

    # Get the header row (first row)
    header = [ws5.cell(row=1, column=col).value for col in range(1, ws5.max_column + 1)]

    return wb5, ws5, countries_stage5, empty_row, header


def insert_separator_column(ws5, max_col):
    max_col += 1
    for row in range(1, ws5.max_row + 1):
        ws5.cell(row=row, column=max_col).fill = PatternFill(start_color="000000", end_color="000000",
                                                             fill_type="solid")
    return max_col


def insert_country_row(ws5, country, countries_stage5):
    total_rooms_row = countries_stage5.get("Total Rooms", ws5.max_row + 1)
    total_camping_row = countries_stage5.get("Total Camping", ws5.max_row + 1)

    # Determine where the country should go
    if "Camping" in country:
        insert_before = total_camping_row  # Camping nationalities should be before "Total Camping"
    else:
        insert_before = total_rooms_row  # Room nationalities should be before "Total Rooms"

    # Find the correct insertion point
    target_row = None
    sorted_countries = sorted((c, r) for c, r in countries_stage5.items() if c and "Total" not in c)

    for existing_country, row in sorted_countries:
        if row >= insert_before:
            break
        if existing_country and country < existing_country:
            target_row = row
            break

    if target_row is None:
        target_row = insert_before  # Default to inserting above "Total Rooms" or "Total Camping"

    ws5.insert_rows(target_row)
    ws5.cell(row=target_row, column=1, value=country)

    # Recalculate index mapping
    new_mapping = {}
    for row in range(2, ws5.max_row + 1):
        cell_value = ws5.cell(row=row, column=1).value
        if cell_value:
            new_mapping[cell_value] = row
    countries_stage5.clear()
    countries_stage5.update(new_mapping)

    return target_row


def copy_cell_styles(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            underline=source_cell.font.underline,
            color=source_cell.font.color
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            start_color=source_cell.fill.start_color.rgb,
            end_color=source_cell.fill.end_color.rgb,
            fill_type=source_cell.fill.fill_type
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )
    target_cell.number_format = source_cell.number_format


def append_stage6_to_stage5(stage5_file, stage6_files, output_file):
    wb5, ws5, countries_stage5, empty_row, header_stage5 = load_stage5_data(stage5_file)

    max_col = ws5.max_column
    first_file_processed = False

    for index, stage6_file in enumerate(stage6_files):
        wb6 = openpyxl.load_workbook(stage6_file)
        ws6 = wb6.active

        # Retrieve the header of Stage 6
        header_stage6 = [ws6.cell(row=1, column=col).value for col in range(1, ws6.max_column + 1)]

        for col, header_value in enumerate(header_stage6, start=max_col + 2):
            ws5.cell(row=1, column=col, value=header_value)

        # Insert separator column and calculate the starting column for the new data
        max_col = insert_separator_column(ws5, max_col)
        start_col = max_col + 1

        for row in range(2, ws6.max_row + 1):
            country = ws6.cell(row=row, column=1).value
            if country is None:
                continue  # Skip None values

            # Find or insert the country row in Stage 5
            if first_file_processed:
                # After the first file, use the first empty row
                target_row = countries_stage5.get(country) or insert_country_row(ws5, country, countries_stage5,
                                                                                 )
            else:
                # In the first file, proceed as before (insert based on the existing order)
                target_row = countries_stage5.get(country) or insert_country_row(ws5, country, countries_stage5,
                                                                                 )

            # Copy the values and styles from Stage 6 to Stage 5
            for col in range(1, ws6.max_column + 1):
                source_cell = ws6.cell(row=row, column=col)
                target_cell = ws5.cell(row=target_row, column=start_col + (col - 1))
                target_cell.value = source_cell.value
                copy_cell_styles(source_cell, target_cell)
                ws5.column_dimensions[
                    openpyxl.utils.get_column_letter(start_col + (col - 1))].width = ws6.column_dimensions.get(
                    openpyxl.utils.get_column_letter(col),
                    ws5.column_dimensions[openpyxl.utils.get_column_letter(start_col + (col - 1))]).width

        # Recalculate the country row indexes after processing the file
        countries_stage5.clear()
        for row in range(2, ws5.max_row + 1):
            cell_value = ws5.cell(row=row, column=1).value
            if cell_value:
                countries_stage5[cell_value] = row

        # Update the `max_col` after processing each Stage 6 file to set the starting column for the next file
        max_col = ws5.max_column

    # Save the output file after all Stage 6 files are processed
    wb5.save(output_file)


def per_nat_stage3(stage5_path, stage6_paths, output_path):
    logger.info(f'Starting with Per Nationality Stage 3')
    append_stage6_to_stage5(stage5_path, stage6_paths, output_path)
    logger.info(f'Per Nationality Stage 3 completed. File saved as {output_path}')


if __name__ == '__main__':
    stage5_path = "per_nat_stage1_output.xlsx"
    stage6_paths = ["per_nat_stage2_output_2024.xlsx",
                    "per_nat_stage2_output_2023.xlsx", ]  # Add all Stage 6 files here
    output_path = "per_nat_stage3_output.xlsx"

    append_stage6_to_stage5(stage5_path, stage6_paths, output_path)
