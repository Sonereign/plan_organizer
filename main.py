import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
import os
import shutil
import threading
from openpyxl import load_workbook
from copy import copy
from datetime import datetime
from stage1 import stage1
from stage2 import stage2
from stage3 import stage3
from stage4 import stage4
from stage5 import stage5


class PlanoKratiseonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Πλάνο Κρατήσεων")
        self.root.geometry("600x600")

        # Variables to store file paths
        self.availability_per_zone_path = None
        self.availability_per_type_path = None
        self.availability_per_nationality_path = None

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        """Create and arrange widgets in the window."""
        # Title Label
        title_label = tk.Label(self.root, text="Πλάνο Κρατήσεων", font=("Arial", 16))
        title_label.pack(pady=10)

        # Availability Per Zone Section
        self.create_file_section("Availability Per Zone", self.availability_per_zone_path, "Select Availability Per Zone File")

        # Availability Per Type Section
        self.create_file_section("Availability Per Type", self.availability_per_type_path, "Select Availability Per Type File")

        # Availability Per Nationality Section
        self.create_file_section("Availability Per Nationality", self.availability_per_nationality_path, "Select Availability Per Nationality File")

        # Process Button
        self.process_button = tk.Button(self.root, text="Process Files", command=self.start_processing, font=("Arial", 12))
        self.process_button.pack(pady=20)

        # Status Label
        self.status_label = tk.Label(self.root, text="", fg="blue")
        self.status_label.pack(pady=10)

    def create_file_section(self, label_text, text_widget, button_text):
        """Helper method to create file selection sections."""
        label = tk.Label(self.root, text=label_text, font=("Arial", 12))
        label.pack(pady=5)

        text_widget = tk.Text(self.root, height=1, width=50)
        text_widget.pack(pady=5)

        button = tk.Button(
            self.root,
            text=button_text,
            command=lambda: self.select_file(label_text, text_widget)
        )
        button.pack(pady=5)

    def select_file(self, file_description, text_widget):
        """Open a file dialog to select an Excel file."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            if file_description == "Availability Per Zone":
                self.availability_per_zone_path = file_path
            elif file_description == "Availability Per Type":
                self.availability_per_type_path = file_path
            elif file_description == "Availability Per Nationality":
                self.availability_per_nationality_path = file_path

            # Update the text widget with the selected file path
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, file_path)
            self.status_label.config(text=f"{file_description} file selected")

    def start_processing(self):
        """Start processing files in a separate thread."""
        if not all([self.availability_per_zone_path, self.availability_per_type_path]):
            messagebox.showerror("Error", "Please select both Availability Per Zone and Availability Per Type files!")
            return

        # Disable the process button to prevent multiple clicks
        self.process_button.config(state=tk.DISABLED)
        self.status_label.config(text="Processing started...")

        # Run processing in a separate thread
        threading.Thread(target=self.process_files, daemon=True).start()

    def process_files(self):
        """Process the selected Excel files using stage1, stage2, stage3, stage4, and stage5."""
        try:
            # Temporary file paths for chaining
            stage1_output = "stage1_output.xlsx"
            stage2_output = "stage2_output.xlsx"
            stage3_output = "stage3_output.xlsx"
            stage4_output = "stage4_output.xlsx"
            stage5_output = "stage5_output.xlsx"

            # Run stages
            self.run_stage(stage1, self.availability_per_zone_path, stage1_output)
            self.run_stage(stage2, stage1_output, self.availability_per_type_path, stage2_output)
            self.run_stage(stage3, stage2_output, stage3_output)
            self.run_stage(stage4, stage3_output, stage4_output)

            # Run stage5 if nationality file is provided
            if self.availability_per_nationality_path:
                self.run_stage(stage5, self.availability_per_nationality_path, stage5_output)

            # Generate final output file name and sheet names
            today = datetime.today().strftime("%d-%m-%y")
            final_output = f"{today}availabilityPerZone&Nationality.xlsx"
            sheet1_name = f"{today}-πληρότητα-units"
            sheet2_name = "εθνικότητες"

            # Create final output by copying sheets from stage4 and stage5 outputs
            self.combine_sheets(stage4_output, stage5_output, final_output, sheet1_name, sheet2_name)

            # Notify user of success
            messagebox.showinfo("Success", f"Final output saved as {final_output}")
            self.status_label.config(text="Processing complete! Final output saved.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
        finally:
            # Re-enable the process button
            self.process_button.config(state=tk.NORMAL)
            # Clean up temporary files
            for file in [stage1_output, stage2_output, stage3_output, stage4_output, stage5_output]:
                if os.path.exists(file):
                    os.remove(file)

    def combine_sheets(self, stage4_output, stage5_output, final_output, sheet1_name, sheet2_name):
        """Combine sheets from stage4 and stage5 outputs into a single Excel file."""
        # Load workbooks
        wb_stage4 = load_workbook(stage4_output)
        wb_stage5 = load_workbook(stage5_output) if self.availability_per_nationality_path else None

        # Create a new workbook for the final output
        wb_final = load_workbook(stage4_output)  # Start with a copy of stage4 output

        # Rename the sheet from stage4 to the custom sheet1 name
        sheet_stage4 = wb_final.active
        sheet_stage4.title = sheet1_name

        # Add sheet from stage5 if available
        if wb_stage5:
            sheet_stage5 = wb_stage5.active
            # Create a new sheet in the final workbook for Stage5 Results
            new_sheet_stage5 = wb_final.create_sheet(sheet2_name)

            # Copy all cells, styles, and formulas from Stage5 sheet to the new sheet
            for row in sheet_stage5.iter_rows():
                for cell in row:
                    new_cell = new_sheet_stage5.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)  # Use copy function
                        new_cell.border = copy(cell.border)  # Use copy function
                        new_cell.fill = copy(cell.fill)  # Use copy function
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)  # Use copy function
                        new_cell.alignment = copy(cell.alignment)  # Use copy function

        # Save the final workbook
        wb_final.save(final_output)

    def run_stage(self, stage_func, *args, stage_name=""):
        """Helper method to run a stage and update the status."""
        self.status_label.config(text=f"Running {stage_name}...")
        self.root.update()
        stage_func(*args)


# Run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = PlanoKratiseonApp(root)
    root.mainloop()