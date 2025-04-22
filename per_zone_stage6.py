from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from logger import logger

def load_stage4(input_file):
    return pd.read_excel(input_file)


def load_stage5_files(stage5_files):
    stage5_data = []
    for file in stage5_files:
        df = pd.read_excel(file)
        stage5_data.append((file, df))
    return stage5_data


def parse_date(date_str):
    try:
        # Remove the day of the week (e.g., "Tue ", "Mon ")
        if isinstance(date_str, str):
            date_str = date_str.split(" ")[-1]  # Keep only the date part

        # Add the year if missing (assuming the year is the current year)
        if len(date_str.split("/")) == 2:  # If the date is in "29/04" format
            current_year = datetime.now().year
            date_str += f"/{current_year}"  # Append the year

        # Parse the date with dayfirst=True
        return pd.to_datetime(date_str, dayfirst=True, errors='coerce')
    except:
        return None


def detect_date_range(file_name, df):
    date_columns = df.columns[2:]
    if len(date_columns) > 0:
        start_date = date_columns[0]
        end_date = date_columns[-1]

        # Ensure the end date is a valid date and not a total column
        if isinstance(end_date, str) and end_date.lower() == "total":
            end_date = date_columns[-2] if len(date_columns) > 1 else "Unknown"

        parsed_start = parse_date(start_date)
        parsed_end = parse_date(end_date)

        logger.info(f"File: {file_name} | Date range: {start_date} to {end_date}")
        return parsed_start, parsed_end
    return None, None


def add_empty_columns(input_file, output_file, start_diff):
    """
    Adds empty columns after the "Capacity" column in the Stage 4 file.
    """
    # Load the workbook using openpyxl
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Find the index of the "Capacity" column
    capacity_col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "Capacity":
            capacity_col_index = col[0].column
            break

    if capacity_col_index is None:
        logger.info("Error: 'Capacity' column not found in the Excel file.")
        return

    # Add empty columns after the "Capacity" column
    if start_diff > 0:
        sheet.insert_cols(capacity_col_index + 1, start_diff)
        logger.info(f"Added {start_diff} empty columns after the Capacity column.")

    # Save the modified workbook
    workbook.save(output_file)
    logger.info(f"Saved modified Stage 4 file with empty columns to {output_file}.")

def add_empty_cells(sheet, row_index, num_empty_cells):
    """
    Adds a specified number of empty cells at index 2 in a given row.
    """
    if num_empty_cells > 0:
        # Insert empty cells at index 2
        for _ in range(num_empty_cells):
            sheet.insert_cols(2)  # Insert a new column at index 2
            # Set the value of the new cell to None (empty)
            for row in sheet.iter_rows(min_row=row_index, max_row=row_index):
                row[1].value = None

def copy_header(from_file, to_file, output_file, num_empty_cells=0, empty_at_index=2):
    """
    Copies the header from the Stage 4 file and writes it into the Stage 5 file as a new row.
    Adds empty cells to the header list at a specific index before writing.

    Args:
        from_file (str): Path to the Stage 4 Excel file.
        to_file (str): Path to the Stage 5 Excel file.
        output_file (str): Path to save the modified Stage 5 file.
        num_empty_cells (int): Number of empty cells to add.
        empty_at_index (int): Index at which to add the empty cells.
    """
    # Step 1: Load the Stage 4 file to extract the header
    stage4_workbook = load_workbook(from_file)
    stage4_sheet = stage4_workbook.active

    # Get the header row from Stage 4 as a list
    stage4_header = [cell.value for cell in stage4_sheet[1]]
    #logger.info("Stage 4 Header (Original):", stage4_header)

    # Step 2: Add empty cells to the header list
    if num_empty_cells > 0:
        for _ in range(num_empty_cells):
            stage4_header.insert(empty_at_index, "")  # Insert empty string at the specified index

    # Step 3: Load the Stage 5 file
    stage5_workbook = load_workbook(to_file)
    stage5_sheet = stage5_workbook.active

    # Step 4: Insert the modified Stage 4 header as a new row in the Stage 5 file
    stage5_sheet.insert_rows(1)  # Insert a new row at the top (position 1)

    # Write the modified Stage 4 header into the new row
    for col_idx, value in enumerate(stage4_header, start=1):
        stage5_sheet.cell(row=1, column=col_idx, value=value)

    # Step 5: Save the modified Stage 5 workbook
    stage5_workbook.save(output_file)
    logger.info(f"Copied header from {from_file} to {to_file}. Added {num_empty_cells} empty cells at index {empty_at_index}. Saved to {output_file}.")


def copy_total_rows_from_stage5(from_file, to_file, output_file, num_empty_cells=2, empty_at_index=2):
    """
    Copies rows starting with "Total Accommodation", "Total Youth Hostel", or "Total Camping"
    from the Stage 5 file and inserts them under their respective counterparts in the Stage 4 file.
    Adds empty cells to the row data at a specific index before writing.
    Skips processing if the Stage 5 file name is 'per_zone_stage5_output_2023.xlsx'.

    Args:
        from_file (str): Path to the Stage 5 Excel file.
        to_file (str): Path to the Stage 4 Excel file.
        output_file (str): Path to save the modified Stage 4 file.
        num_empty_cells (int): Number of empty cells to add.
        empty_at_index (int): Index at which to add the empty cells.
    """

    # Step 2: Load the Stage 5 file to extract the total rows
    stage5_df = pd.read_excel(from_file)

    # Identify the rows to copy (check if the first column starts with the keywords)
    total_rows = stage5_df[
        stage5_df.iloc[:, 0].str.startswith("Total Accommodation") |
        stage5_df.iloc[:, 0].str.startswith("Total Youth Hostel") |
        stage5_df.iloc[:, 0].str.startswith("Total Camping")
    ]

    if total_rows.empty:
        logger.info(f"Warning: No total rows found in the Stage 5 file {from_file}.")
        return

    # Step 3: Load the Stage 4 file using openpyxl
    workbook = load_workbook(to_file)
    sheet = workbook.active

    # Step 4: Iterate through the total rows from Stage 5
    for idx, row in total_rows.iterrows():
        # Extract the keyword from the row (e.g., "Total Accommodation")
        keyword = next(
            (k for k in ["Total Accommodation", "Total Youth Hostel", "Total Camping"]
            if str(row.iloc[0]).startswith(k)),
            None
        )

        if not keyword:
            continue  # Skip if no matching keyword is found

        # Step 5: Find the corresponding row in Stage 4
        target_row_index = None
        for row_idx, sheet_row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1), start=1):
            if sheet_row[0].value and str(sheet_row[0].value).startswith(keyword):
                target_row_index = row_idx
                break

        if target_row_index is None:
            logger.info(f"Warning: No matching row found in Stage 4 for '{keyword}'.")
            continue

        # Step 6: Convert the row to a list
        row_list = row.tolist()

        # Step 7: Add empty cells to the row list
        if num_empty_cells > 0:
            for _ in range(num_empty_cells):
                row_list.insert(empty_at_index, "")  # Insert empty string at the specified index

        # Step 8: Insert the row from Stage 5 below the target row in Stage 4
        sheet.insert_rows(target_row_index + 1)  # Insert a new row below the target row
        for col_idx, value in enumerate(row_list, start=1):
            sheet.cell(row=target_row_index + 1, column=col_idx, value=value)

    # Step 9: Save the modified workbook
    workbook.save(output_file)
    logger.info(f"Copied total rows from {from_file} to Stage 4 file. Added {num_empty_cells} empty cells at index {empty_at_index}. Saved to {output_file}.")

def calculate_days_difference(date1, date2):
    """
    Calculates the absolute difference in days between two dates, ignoring the year.
    """
    # Extract month and day from the dates
    date1_md = (date1.month, date1.day)
    date2_md = (date2.month, date2.day)

    # Create datetime objects for the same year (year is irrelevant)
    current_year = datetime.now().year
    date1_fixed = datetime(current_year, date1_md[0], date1_md[1])
    date2_fixed = datetime(current_year, date2_md[0], date2_md[1])

    # Calculate the absolute difference in days
    return abs((date1_fixed - date2_fixed).days)

def extract_month_day(date):
    """
    Extracts the month and day from a date and returns a tuple (month, day).
    """
    return (date.month, date.day)

def calculate_and_print_date_differences(stage4_start, stage4_end, stage5_files):
    """
    Calculates the absolute days difference between the earliest starting date
    and latest ending date in the Stage 5 files compared to the starting and ending dates
    of each file. Ignores the year when determining the earliest and latest dates.

    Args:
        stage4_start (datetime): The starting date of the Stage 4 file.
        stage4_end (datetime): The ending date of the Stage 4 file.
        stage5_files (list): A list of file paths to Stage 5 Excel files.

    Returns:
        list: A list of dictionaries containing the date differences for Stage 4 and each Stage 5 file.
              Example:
              [
                  {"file": "Stage4", "start_date": "04-29", "end_date": "04-30", "start_diff": 2, "end_diff": 2},
                  {"file": "stage5_file1.xlsx", "start_date": "04-27", "end_date": "05-02", "start_diff": 0, "end_diff": 0},
                  {"file": "stage5_file2.xlsx", "start_date": "04-28", "end_date": "05-01", "start_diff": 1, "end_diff": 1}
              ]
    """
    results = []

    # Load all Stage 5 files and collect their starting and ending dates
    all_start_dates = []
    all_end_dates = []

    for file in stage5_files:
        df = pd.read_excel(file)
        start_date, end_date = detect_date_range(file, df)
        if start_date:
            all_start_dates.append(start_date)
        if end_date:
            all_end_dates.append(end_date)

    if not all_start_dates or not all_end_dates:
        logger.info("Error: Could not determine valid date ranges for Stage 5 files.")
        return results

    # Extract month and day from all starting and ending dates
    all_start_md = [extract_month_day(date) for date in all_start_dates]
    all_end_md = [extract_month_day(date) for date in all_end_dates]

    # Find the earliest starting date and latest ending date (ignoring the year)
    earliest_start_md = min(all_start_md)
    latest_end_md = max(all_end_md)

    # Convert back to datetime objects for the current year (year is irrelevant)
    current_year = datetime.now().year
    earliest_start = datetime(current_year, earliest_start_md[0], earliest_start_md[1])
    latest_end = datetime(current_year, latest_end_md[0], latest_end_md[1])

    # Calculate and store the results for Stage 4
    stage4_start_diff = calculate_days_difference(stage4_start, earliest_start)
    stage4_end_diff = calculate_days_difference(stage4_end, latest_end)
    results.append({
        "file": "Stage4",
        "start_date": stage4_start.strftime("%m-%d"),
        "end_date": stage4_end.strftime("%m-%d"),
        "start_diff": stage4_start_diff,
        "end_diff": stage4_end_diff
    })

    # Calculate and store the results for each Stage 5 file
    for file in stage5_files:
        df = pd.read_excel(file)
        file_start, file_end = detect_date_range(file, df)
        if file_start and file_end:
            start_diff = calculate_days_difference(file_start, earliest_start)
            end_diff = calculate_days_difference(file_end, latest_end)
            results.append({
                "file": file,
                "start_date": file_start.strftime("%m-%d"),
                "end_date": file_end.strftime("%m-%d"),
                "start_diff": start_diff,
                "end_diff": end_diff
            })

    return results

def per_zone_stage6(input_file, stage5_files, output_file):
    # Load the Stage 4 file
    stage4_df = load_stage4(input_file)
    stage4_start, stage4_end = detect_date_range(input_file, stage4_df)

    if stage4_start is None or stage4_end is None:
        logger.info(f"Warning: Could not determine date range for Stage 4 file {input_file}")

    # Calculate and print date differences
    date_differences = calculate_and_print_date_differences(stage4_start, stage4_end, stage5_files)

    # Load the Stage 5 files
    stage5_data = load_stage5_files(stage5_files)
    all_start_dates = []
    all_end_dates = []

    for file_name, df in stage5_data:
        start_date, end_date = detect_date_range(file_name, df)
        if start_date:
            all_start_dates.append(start_date)
        if end_date:
            all_end_dates.append(end_date)

    if all_start_dates and all_end_dates and stage4_start and stage4_end:
        # Extract month and day from all starting and ending dates
        all_start_md = [extract_month_day(date) for date in all_start_dates]
        all_end_md = [extract_month_day(date) for date in all_end_dates]

        # Find the earliest starting date and latest ending date (ignoring the year)
        earliest_start_md = min(all_start_md)
        latest_end_md = max(all_end_md)

        # Convert back to datetime objects for the current year (year is irrelevant)
        current_year = datetime.now().year
        earliest_start = datetime(current_year, earliest_start_md[0], earliest_start_md[1])
        latest_end = datetime(current_year, latest_end_md[0], latest_end_md[1])

        # Calculate the days difference (ignoring the year)
        start_diff = calculate_days_difference(stage4_start, earliest_start)
        end_diff = calculate_days_difference(stage4_end, latest_end)

        # Add empty columns to Stage 4
        add_empty_columns(input_file, output_file, start_diff)

        num_empty_cells = 0

        # Iterate through all Stage 5 files
        for stage5_file in reversed(stage5_files):
            # Extract start_diff for a specific file
            for result in date_differences:
                if result["file"] == stage5_file:
                    num_empty_cells = result["start_diff"]
                    logger.info(f"Start Diff for {stage5_file=}: {start_diff}")
                    break

            # Copy header from Stage 5 to Stage 4
            copy_header(stage5_file, output_file, output_file, num_empty_cells=num_empty_cells)

            # Copy total rows from Stage 5 to Stage 4
            copy_total_rows_from_stage5(stage5_file, output_file, output_file, num_empty_cells=num_empty_cells)
    else:
        logger.info("Error: Could not determine valid date ranges for comparison.")

    # Placeholder for further processing
    logger.info("Loaded Stage 4 and Stage 5 files successfully")


if __name__ == "__main__":
    INPUT_FILE = "per_zone_stage4_output.xlsx"
    STAGE5_FILES = ["per_zone_stage5_output_2024.xlsx", "per_zone_stage5_output_2023.xlsx"]
    OUTPUT_FILE = "per_zone_stage6_output.xlsx"
    per_zone_stage6(INPUT_FILE, STAGE5_FILES, OUTPUT_FILE)