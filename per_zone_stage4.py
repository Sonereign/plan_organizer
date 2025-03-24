from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from logger import logger

# Hardcoded capacities for accommodations
ACCOMMODATION_CAPACITIES = {
    "APT": 2,
    "Beach": 3,
    "for2": 7,
    "for5": 14,
    "for6": 5,
    ".LUX for 4": 51,
    ".Safari Tent 5pax": 12,
    ".Sea Safari 4pax": 31,
    ".Skyline 3pax": 11,
    ".Standard Mobile Home": 31,
    ".ΤΡΟΧΟΣΠΙΤΑ DELUXE": 8,
    ".ΤΡΟΧΟΣΠΙΤΑ SEA VIEW": 24,
    ".ΤΡΟΧΟΣΠΙΤΑ standard": 8
}

# Hardcoded capacities for camping areas
CAMPING_CAPACITIES = {
    "1": 0,
    "2": 20,
    "3": 81,
    "4": 23,
    "5": 44,
    "6": 11,
    "7": 32,
    "Z": 27,  # English Z
    "Ζ": 27,  # Greek Zeta
    "K": 80,  # English K
    "Κ": 80,  # Greek Kappa
    "Δ": 12,
    "Ε": 14,
    "Ι": 4
}

# Constants for formatting
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))


def detect_groups(df):
    """Detect groups based on empty rows."""
    groups = []
    current_group = {"start_row": None, "types": [],  "name": ""}
    first_group = True
    second_group = True

    # Start from row 2 (index 1) to skip the header
    for index, row in df.iloc[1:].iterrows():
        # Check if the row is empty (all values are NaN)
        if row.isnull().all():
            # If a group has been started, finalize it
            if current_group["start_row"] is not None:
                current_group["end_row"] = index + 1  # Excel rows are 1-based
                if first_group:
                    current_group["name"] = "Accommodations"
                    first_group = False
                    second_group = True
                elif second_group:
                    current_group["name"] = "Youth Hostel"
                    second_group = False
                groups.append(current_group)
                current_group = {"start_row": None, "types": [], "name": ""}
        else:
            # If no group has been started, start a new one
            if current_group["start_row"] is None:
                current_group["start_row"] = index + 1  # Excel rows are 1-based
            # Add the accommodation type (first column value)
            current_group["types"].append(row[0])

    # Add the last group if it exists
    if current_group["start_row"] is not None:
        current_group["end_row"] = len(df)
        current_group["name"] = "Camping"
        groups.append(current_group)

    print(f'{groups}')
    return groups


def add_totals_and_occupancy_rows(df, group, is_last_group=False):
    """Add 'Totals' and 'Πληρότητα' rows with Excel formulas."""
    start_row = group["start_row"] - 1  # Convert to 0-based index
    end_row = group["end_row"] - 1  # Convert to 0-based index
    current_year = datetime.now().year

    # Create a "Totals" row
    totals_row = [f'Total {group["name"]} {current_year}', ""]  # Initialize with "Totals" and empty capacity

    # Add SUM formula for the Capacity column (column B)
    capacity_col_letter = get_column_letter(2)  # Column B
    if not is_last_group:
        # capacity_formula = f"=SUM({capacity_col_letter}{start_row + 1}:{capacity_col_letter}{end_row})"
        capacity_formula = ""
    else:
        # capacity_formula = f"=SUM({capacity_col_letter}{start_row + 1}:{capacity_col_letter}{end_row + 1})"
        capacity_formula = ""

    totals_row[1] = capacity_formula  # Add the formula to the Capacity column

    # Add SUM formulas for each day's column
    for col_index in range(2, len(df.columns)):
        # Get the column letter (e.g., 2 -> 'C', 3 -> 'D')
        col_letter = get_column_letter(col_index + 1)  # +1 because Excel columns are 1-based

        # Create the SUM formula for the column
        if not is_last_group:
            # formula = f"=SUM({col_letter}{start_row + 1}:{col_letter}{end_row})"
            formula = ""
            totals_row.append(formula)
        else:
            # formula = f"=SUM({col_letter}{start_row + 1}:{col_letter}{end_row + 1})"
            formula = ""
            totals_row.append(formula)

    # Insert the "Totals" row
    if is_last_group:
        # For the last group, insert the "Totals" row after the last non-empty row
        df.loc[end_row + 0.5] = totals_row
    else:
        # For other groups, insert the "Totals" row before the empty line
        df.loc[end_row - 0.5] = totals_row

    # Reindex the DataFrame
    df = df.sort_index().reset_index(drop=True)

    # Add a "Πληρότητα" row after the "Totals" row
    occupancy_row = ["Πληρότητα", ""]  # Initialize with "Πληρότητα" and empty capacity

    # Add occupancy percentage formulas for each day's column
    for col_index in range(2, len(df.columns)):
        # Get the column letter (e.g., 2 -> 'C', 3 -> 'D')
        col_letter = get_column_letter(col_index + 1)  # +1 because Excel columns are 1-based

        # Create the occupancy percentage formula for the column
        if not is_last_group:
            # formula = f"=({col_letter}{end_row + 1}/{capacity_col_letter}{end_row + 1})"
            formula = ""
            occupancy_row.append(formula)
        else:
            # formula = f"=({col_letter}{end_row + 2}/{capacity_col_letter}{end_row + 2})"
            formula = ""
            occupancy_row.append(formula)

    # Insert the "Πληρότητα" row
    if is_last_group:
        # For the last group, insert the "Πληρότητα" row after the "Totals" row
        df.loc[end_row + 1.5] = occupancy_row
    else:
        # For other groups, insert the "Πληρότητα" row after the "Totals" row
        df.loc[end_row + 0.5] = occupancy_row

    # Reindex the DataFrame
    df = df.sort_index().reset_index(drop=True)

    return df


def per_zone_stage4(input_file, output_file):
    """
    Process the input file (output of stage3) and save the result to the output file.
    """
    logger.info("#######################################################")
    logger.info(f"Running Per Zone Stage 4 with {input_file=}")
    # Load the Excel file
    df = pd.read_excel(input_file, sheet_name='Sheet1', header=None)

    # Detect groups dynamically
    groups = detect_groups(df)

    # Add "Totals" and "Πληρότητα" rows for each group
    for i, group in enumerate(groups):
        is_last_group = (i == len(groups) - 1)  # Check if this is the last group
        df = add_totals_and_occupancy_rows(df, group, is_last_group)
        # Recalculate row numbers for subsequent groups
        for j in range(i + 1, len(groups)):
            groups[j]["start_row"] += 2  # Adjust for both "Totals" and "Πληρότητα" rows
            groups[j]["end_row"] += 2

    # Add a new column for the sum of each row
    sum_col_index = len(df.columns)  # Index of the new sum column
    df[sum_col_index] = ""  # Initialize the new column with empty values

    # Name the new column "Total" in the first row
    df.at[0, sum_col_index] = "Total"

    # Add formulas for the sum column (skip the first row and "Πληρότητα" row)
    for index, row in df.iterrows():
        if index > 0 and not pd.isna(row[0]) and row[0] != "Πληρότητα":  # Skip header row, empty rows, and "Πληρότητα" row
            # Get the range of date columns (columns 2 to sum_col_index - 1)
            start_col = get_column_letter(3)  # Column C (first date column)
            end_col = get_column_letter(sum_col_index)  # Last date column
            # Add the SUM formula for the row
            # formula = f"=SUM({start_col}{index + 1}:{end_col}{index + 1})"
            formula = ""
            df.at[index, sum_col_index] = formula

    # Save the updated DataFrame to a new Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Stage4 Results")

        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = workbook["Stage4 Results"]

        # Ensure the sheet is visible
        worksheet.sheet_state = 'visible'

        # Set the width of the first column to 22
        worksheet.column_dimensions['A'].width = 22

        # Define styles
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue for headers
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply styles to the "Total" column and rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=sum_col_index + 1, max_col=sum_col_index + 1):
            for cell in row:
                cell.fill = yellow_fill
                cell.font = bold_font
                cell.border = thin_border

        # Apply styles to the "Totals" rows
        for group in groups:
            totals_row_index = group["end_row"] + 1 if group == groups[-1] else group["end_row"]
            for row in worksheet.iter_rows(min_row=totals_row_index, max_row=totals_row_index, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.fill = yellow_fill
                    cell.font = bold_font
                    cell.border = thin_border

        # Apply styles to the "Πληρότητα" rows
        for group in groups:
            occupancy_row_index = group["end_row"] + 2 if group == groups[-1] else group["end_row"] + 1
            for row in worksheet.iter_rows(min_row=occupancy_row_index, max_row=occupancy_row_index, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.fill = light_green_fill
                    cell.font = bold_font
                    cell.border = thin_border
                    if cell.column > 2:  # Apply percentage format to columns after Capacity
                        cell.number_format = "0.00%"

        # Apply styles to the header cells (only for "Fri", "Sat", "Sun")
        for col in range(3, worksheet.max_column):  # Start from column C (Day 1)
            header_cell = worksheet.cell(row=1, column=col)
            if isinstance(header_cell.value, str) and any(day in header_cell.value for day in ["Fri", "Sat", "Sun"]):
                header_cell.fill = header_fill
                header_cell.font = bold_font
                header_cell.border = thin_border

        # Freeze pane at B2
        worksheet.freeze_panes = "C2"

    logger.info(f"Per Zone Stage 4 completed. File saved as {output_file}")


if __name__ == "__main__":
    # Default file paths (for standalone execution)
    INPUT_FILE = "per_zone_stage3_output.xlsx"
    OUTPUT_FILE = "per_zone_stage4_output.xlsx"

    # Run stage4
    per_zone_stage4(INPUT_FILE, OUTPUT_FILE)