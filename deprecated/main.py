import os
import tkinter as tk
from copy import copy
from tkinter import messagebox, filedialog
import threading
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from per_zone_stage1 import per_zone_stage1
from per_nat_stage6 import per_nat_stage6
from per_zone_stage2 import per_zone_stage2
from per_zone_stage3 import per_zone_stage3
from per_zone_stage4 import per_zone_stage4
from per_nat_stage1 import per_nat_stage1
from per_nat_stage2 import per_nat_stage2
from per_nat_stage3 import per_nat_stage3
from per_nat_stage4 import per_nat_stage4
from per_nat_stage5 import per_nat_stage5

CLEANUP_OUTPUTS = False


class PlanoKratiseonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Πλάνο Κρατήσεων")
        self.root.geometry("600x700")
        self.root.configure(bg="#f0f0f0")

        # Variables to store file paths
        self.availability_per_zone_path = None
        self.availability_per_type_path = None
        self.availability_per_nationality_path = None
        self.previous_years_paths = {}

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        title_label = tk.Label(self.root, text="Πλάνο Κρατήσεων", font=("Arial", 18, "bold"), bg="#f0f0f0")
        title_label.pack(pady=10)

        file_frame = tk.Frame(self.root, bg="#f0f0f0")
        file_frame.pack(pady=10, fill="x", padx=20)

        self.availability_per_zone_text = self.create_file_section(file_frame, "Availability Per Zone")
        self.availability_per_type_text = self.create_file_section(file_frame, "Availability Per Type")

        nationality_frame = tk.LabelFrame(file_frame, text="Availability Per Nationality & Previous Years",
                                          font=("Arial", 12, "bold"), bg="#f0f0f0", padx=10, pady=10)
        nationality_frame.pack(fill="x", pady=10)

        self.availability_per_nationality_text = self.create_file_section(nationality_frame, "Current Year")
        self.previous_years_frame = tk.Frame(nationality_frame, bg="#f0f0f0")
        self.previous_years_frame.pack(fill="x", pady=5)

        add_year_button = tk.Button(
            nationality_frame, text="+ Add Year", command=self.add_previous_year,
            font=("Arial", 10), bg="#008CBA", fg="white"
        )
        add_year_button.pack(pady=5)

        self.process_button = tk.Button(
            self.root, text="Process Files", command=self.start_processing,
            font=("Arial", 12), bg="#4CAF50", fg="white", padx=10, pady=5
        )
        self.process_button.pack(pady=20)

        # Cleanup Checkbox
        self.cleanup_var = tk.BooleanVar(value=True)
        self.cleanup_checkbox = tk.Checkbutton(root, text="Enable Cleanup", variable=self.cleanup_var,
                                               command=self.toggle_cleanup)
        self.cleanup_checkbox.pack()

        # Status Label
        self.status_label = tk.Label(self.root, text="", fg="blue", bg="#f0f0f0", font=("Arial", 10))
        self.status_label.pack(pady=10)

    def create_file_section(self, parent, label_text):
        frame = tk.Frame(parent, bg="#f0f0f0")
        frame.pack(fill="x", pady=5)

        label = tk.Label(frame, text=label_text, font=("Arial", 12), bg="#f0f0f0")
        label.pack(side="left", padx=5)

        text_widget = tk.Entry(frame, width=40, font=("Arial", 10))
        text_widget.pack(side="left", padx=5, pady=5)

        button = tk.Button(
            frame, text="Browse", command=lambda: self.select_file(label_text, text_widget),
            font=("Arial", 10), bg="#008CBA", fg="white"
        )
        button.pack(side="right", padx=5)

        return text_widget

    def add_previous_year(self):
        year = 2024 - len(self.previous_years_paths)
        frame = tk.Frame(self.previous_years_frame, bg="#f0f0f0")
        frame.pack(fill="x", pady=2)

        label = tk.Label(frame, text=f"Year {year}", font=("Arial", 12), bg="#f0f0f0")
        label.pack(side="left", padx=5)

        text_widget = tk.Entry(frame, width=35, font=("Arial", 10))
        text_widget.pack(side="left", padx=5, pady=5)

        button = tk.Button(
            frame, text="Browse", command=lambda: self.select_file(f"Year {year}", text_widget),
            font=("Arial", 10), bg="#008CBA", fg="white"
        )
        button.pack(side="right", padx=5)

        self.previous_years_paths[year] = text_widget

    def select_file(self, file_description, text_widget):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            text_widget.delete(0, tk.END)
            text_widget.insert(0, file_path)

            if file_description == "Availability Per Zone":
                self.availability_per_zone_path = file_path
            elif file_description == "Availability Per Type":
                self.availability_per_type_path = file_path
            elif file_description == "Current Year":
                self.availability_per_nationality_path = file_path
            elif "Year" in file_description:
                year = int(file_description.split()[1])
                self.previous_years_paths[year] = file_path

    def start_processing(self):
        if not all([self.availability_per_zone_path, self.availability_per_type_path]):
            messagebox.showerror("Error", "Please select both Availability Per Zone and Availability Per Type files!")
            return

        self.process_button.config(state=tk.DISABLED)
        threading.Thread(target=self.process_files, daemon=True).start()

    def process_files(self):
        """Process the selected Excel files using various stages."""
        try:
            # Temporary file paths for chaining
            stage1_output = "stage1_output.xlsx"
            stage2_output = "stage2_output.xlsx"
            stage3_output = "stage3_output.xlsx"
            stage4_output = "stage4_output.xlsx"
            stage5_output = "stage5_output.xlsx"
            stage6_output_filenames = []
            stage7_output = "stage7_output.xlsx"
            stage8_output = "stage8_output.xlsx"
            stage9_output = "stage9_output.xlsx"
            stage10_output = "stage10_output.xlsx"

            # Run stages
            self.run_stage(per_zone_stage1, self.availability_per_zone_path, stage1_output)
            self.run_stage(per_zone_stage2, stage1_output, self.availability_per_type_path, stage2_output)
            self.run_stage(per_zone_stage3, stage2_output, stage3_output)
            self.run_stage(per_zone_stage4, stage3_output, stage4_output)

            # Run stage5 if nationality file is provided
            if self.availability_per_nationality_path:
                self.run_stage(per_nat_stage1, self.availability_per_nationality_path, stage5_output)

            # Run Stage 6 for previous years
            for year, file_path in self.previous_years_paths.items():
                output_file = f"stage6_Output_{year}.xlsx"
                stage6_output_filenames.append(output_file)  # Append file name to list
                per_nat_stage2(input_file=file_path, output_file=output_file, year=year)

            previous_years = list(self.previous_years_paths.keys())  # Extract years from dictionary keys
            number_of_previous_year_data = len(stage6_output_filenames)

            # Generate final output file name and sheet names
            today = datetime.today().strftime("%d-%m-%y")
            final_output = f"{today}_availabilityPerZone&Nationality.xlsx"
            sheet1_name = f"{today}-πληρότητα-units"
            sheet2_name = "εθνικότητες"

            if not stage6_output_filenames:
                # Create final output by combining sheets from stage4 and stage5 outputs
                self.combine_sheets(stage4_output, stage5_output, final_output, sheet1_name, sheet2_name)
            else:
                # Create final output by combining sheets from stage4 and stage10 outputs
                self.run_stage(per_nat_stage3, stage5_output, stage6_output_filenames, stage7_output)
                self.run_stage(per_nat_stage4, stage7_output, stage8_output, previous_years, number_of_previous_year_data)
                self.run_stage(per_nat_stage5, stage8_output, stage9_output, previous_years)
                self.run_stage(per_nat_stage6, stage9_output, stage10_output, previous_years)
                self.combine_sheets(stage4_output, stage10_output, final_output, sheet1_name, sheet2_name)

            # Notify user of success
            self.status_label.config(text="Processing complete!")
            messagebox.showinfo("Success", f"Final output saved as {final_output}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
        finally:
            self.process_button.config(state=tk.NORMAL)
            if CLEANUP_OUTPUTS:
                # Clean up temporary files
                for file in [stage1_output, stage2_output, stage3_output, stage4_output, stage5_output, stage7_output,
                             stage8_output, stage9_output, stage10_output]:
                    if os.path.exists(file):
                        os.remove(file)
                for file in stage6_output_filenames:
                    if os.path.exists(file):
                        os.remove(file)

    def combine_sheets(self, stage4_output, stage5_output, final_output, sheet1_name, sheet2_name):
        """Combine sheets from stage4 and stage5 outputs into a single Excel file."""
        # Load workbooks
        wb_stage4 = load_workbook(stage4_output)
        wb_stage5 = load_workbook(stage5_output) if self.availability_per_nationality_path else None

        # Create a new workbook for the final output
        wb_final = load_workbook(stage4_output)  # Start with a copy of stage4 output

        # Rename the sheet from stage4 to the custom sheet1 name
        sheet_stage4 = wb_final.active
        sheet_stage4.title = sheet1_name

        # Add sheet from stage5 if available
        if wb_stage5:
            sheet_stage5 = wb_stage5.active
            # Create a new sheet in the final workbook for Stage5 Results
            new_sheet_stage5 = wb_final.create_sheet(sheet2_name)

            # Copy all cells, styles, and formulas from Stage5 sheet to the new sheet
            for row in sheet_stage5.iter_rows():
                for cell in row:
                    new_cell = new_sheet_stage5.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)  # Use copy function
                        new_cell.border = copy(cell.border)  # Use copy function
                        new_cell.fill = copy(cell.fill)  # Use copy function
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)  # Use copy function
                        new_cell.alignment = copy(cell.alignment)  # Use copy function

        # Save the final workbook
        wb_final.save(final_output)
        self.apply_conditional_formatting(final_output)

    def apply_conditional_formatting(self, file_path):
        # Load the workbook and select the specified worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb["εθνικότητες"]

        # Find columns that start with "Percent difference"
        header_row = ws[1]  # Assuming headers are in the first row
        percent_diff_cols = [cell.column for cell in header_row if
                             cell.value and cell.value.startswith("Percent difference")]

        # Apply conditional formatting to each identified column
        for percent_diff_col in percent_diff_cols:
            percent_diff_range = f"{get_column_letter(percent_diff_col)}2:{get_column_letter(percent_diff_col)}{ws.max_row}"
            color_scale_rule = ColorScaleRule(
                start_type="num", start_value=-1, start_color="FFCCCC",  # Red for negative
                mid_type="num", mid_value=0, mid_color="FFFFFF",  # White for neutral
                end_type="num", end_value=1, end_color="CCFFCC"  # Green for positive
            )
            ws.conditional_formatting.add(percent_diff_range, color_scale_rule)

        # Save the workbook
        wb.save(file_path)
        print("Conditional formatting applied successfully!")

    def run_stage(self, stage_func, *args, stage_name=""):
        """Helper method to run a stage and update the status."""
        try:
            self.status_label.config(text=f"Running {stage_name}...")
            self.root.update()
            stage_func(*args)
        except Exception as e:
            messagebox.showerror("Error", f"Error in {stage_name}: {e}")

    # Function to handle cleanup checkbox state
    def toggle_cleanup(self):
        global CLEANUP_OUTPUTS
        if not CLEANUP_OUTPUTS:
            CLEANUP_OUTPUTS = True
            print("Temporary files will be removed.")
        else:
            CLEANUP_OUTPUTS = False
            print("Temporary files will stay.")


# Run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = PlanoKratiseonApp(root)
    root.mainloop()
