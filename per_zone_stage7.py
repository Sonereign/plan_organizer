import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from logger import logger


def detect_breakpoints(ws):
    """Identify key breakpoints in the worksheet"""
    logger.info("🔍 Detecting breakpoints...")
    breakpoints = {
        'last_category_row': None,
        'first_empty_row': None,
        'final_empty_row': None
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Check for Category rows
        if row and row[0] == "Category":
            breakpoints['last_category_row'] = row_idx
            logger.info(f"  • Category row found at {row_idx}")

        # Check for empty rows
        if row and all(cell is None for cell in row):
            if breakpoints['first_empty_row'] is None:
                breakpoints['first_empty_row'] = row_idx
                logger.info(f"  • First empty row at {row_idx}")
            breakpoints['final_empty_row'] = row_idx

    logger.info("📌 Breakpoints identified:")
    for name, row in breakpoints.items():
        logger.info(f"  {name.replace('_', ' ').title()}: {row or 'Not found'}")

    return breakpoints


def locate_target_rows(ws):
    """Find all target sum rows and their related rows"""
    logger.info("🔍 Locating target rows...")
    current_year = datetime.datetime.now().year
    target_data = {
        f'Total Accommodations {current_year}': {
            'sum_row': None,
            'capacity_row': None,
            'capacity_value': 0,
            'occupancy_row': None
        },
        f'Total Youth Hostel {current_year}': {
            'sum_row': None,
            'capacity_row': None,
            'capacity_value': 0,
            'occupancy_row': None
        },
        f'Total Camping {current_year}': {
            'sum_row': None,
            'capacity_row': None,
            'capacity_value': 0,
            'occupancy_row': None
        }
    }

    # First pass: Find sum rows and capacity rows
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] in target_data:
            target_data[row[0]]['sum_row'] = row_idx
            target_data[row[0]]['capacity_row'] = row_idx
            logger.info(f"  • Found {row[0]} at row {row_idx} (capacity at row {row_idx})")

    # Second pass: Find Πληρότητα rows (they come in fixed order: Accommodations, Youth Hostel, Camping)
    occupancy_rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == 'Πληρότητα':
            occupancy_rows.append(row_idx)

    # Assign Πληρότητα rows in fixed order
    if len(occupancy_rows) >= 3:
        target_data[f'Total Accommodations {current_year}']['occupancy_row'] = occupancy_rows[0]
        target_data[f'Total Youth Hostel {current_year}']['occupancy_row'] = occupancy_rows[1]
        target_data[f'Total Camping {current_year}']['occupancy_row'] = occupancy_rows[2]
        logger.info(
            f"  • Found Πληρότητα rows at {occupancy_rows[0]} (Accommodations), {occupancy_rows[1]} (Youth Hostel), {occupancy_rows[2]} (Camping)")
    else:
        logger.info("⚠️ Warning: Not enough Πληρότητα rows found (need 3)")

    return target_data


def calculate_total_capacity(ws, start_row, end_row):
    """Calculate total capacity between two rows in column B"""
    total_capacity = 0
    for row_idx in range(start_row, end_row):
        capacity = ws.cell(row=row_idx, column=2).value  # Column B is capacity
        if isinstance(capacity, (int, float)):
            total_capacity += capacity
    return total_capacity


def determine_stop_row(ws, current_row, breakpoints):
    """Find where the summation should stop (going upward)"""
    for row_above in range(current_row - 1, 0, -1):
        # Check if we hit a category row
        if ws.cell(row=row_above, column=1).value == "Category":
            logger.info(f"    Stop condition: Category row at {row_above}")
            return row_above + 1

        # Check if we hit a completely empty row
        if all(ws.cell(row=row_above, column=c).value is None
               for c in range(1, ws.max_column + 1)):
            logger.info(f"    Stop condition: Empty row at {row_above}")
            return row_above + 1

    logger.info("    Warning: No breakpoint found, defaulting to row 2")
    return 2  # Default fallback


def calculate_occupancy_rates(ws, target_data):
    """Calculate and write occupancy rates (Πληρότητα) for each date"""
    logger.info("📈 Calculating occupancy rates...")
    current_year = datetime.datetime.now().year

    categories = [f'Total Accommodations {current_year}', f'Total Youth Hostel {current_year}',
                  f'Total Camping {current_year}']
    for category in categories:
        data = target_data[category]
        if not data.get('occupancy_row'):
            logger.info(f"⚠️ Warning: No Πληρότητα row found for {category}")
            continue

        sum_row = data['sum_row']
        capacity_row = data['capacity_row']
        occupancy_row = data['occupancy_row']

        if not capacity_row:
            logger.info(f"⚠️ Warning: No capacity row found for {category}")
            continue

        logger.info(f"  Processing Πληρότητα for {category} at row {occupancy_row}")
        logger.info(f"    Using sum row {sum_row} and capacity row {capacity_row}")

        # Calculate occupancy for each date column (columns C onward)
        for col_idx in range(3, ws.max_column):
            col_letter = get_column_letter(col_idx)

            # Create formula: =sum_cell/capacity_cell
            sum_cell = f"{col_letter}{sum_row}"
            capacity_cell = f"B{capacity_row}"  # Capacity is always in column B
            formula = f"={sum_cell}/{capacity_cell}"
            ws[f"{col_letter}{occupancy_row}"] = formula

            # Format as percentage
            ws[f"{col_letter}{occupancy_row}"].number_format = '0.00%'

            # Debug output for first and last columns
            if col_idx == 3 or col_idx == ws.max_column:
                logger.info(f"    {col_letter}: {formula}")


def process_category(ws, category_name, data, breakpoints):
    """Process a single category (sums and capacity)"""
    sum_row = data['sum_row']
    capacity_row = data['capacity_row']

    logger.info(f"  Processing {category_name}:")
    logger.info(f"    Sum row: {sum_row}, Capacity row: {capacity_row}")

    # Determine where to stop summing (going upward from sum_row)
    stop_row = determine_stop_row(ws, sum_row, breakpoints)

    # Calculate total capacity (column B)
    data['capacity_value'] = calculate_total_capacity(ws, stop_row, sum_row)
    logger.info(f"    Calculated capacity: {data['capacity_value']} (rows {stop_row}-{sum_row - 1})")

    # Write capacity to the next row (column B)
    ws.cell(row=capacity_row, column=2).value = data['capacity_value']

    # Apply sum formulas to each data column (columns C onward)
    for col_idx in range(3, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        formula_range = f"{col_letter}{stop_row}:{col_letter}{sum_row - 1}"
        ws[f"{col_letter}{sum_row}"] = f"=SUM({formula_range})"

        # Debug output for first and last columns
        if col_idx == 3 or col_idx == ws.max_column:
            logger.info(f"    {col_letter}: Summing rows {stop_row}-{sum_row - 1}")


def calculate_total_column(ws):
    """Calculate and write the Total column for all rows except Πληρότητα"""
    logger.info("🧮 Calculating Total column for all relevant rows...")

    # Find the last column (Total column)
    last_col = ws.max_column
    total_col_letter = get_column_letter(last_col)
    logger.info(f"  Total column found at {total_col_letter}")

    # Process all rows in the worksheet
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        # Skip header rows and empty rows
        if row[0].value in ["Category", "Capacity"] or all(cell.value is None for cell in row):
            continue

        # Skip Πληρότητα rows
        if row[0].value == 'Πληρότητα':
            continue

        # Skip if already has a value in total column
        if row[-1].value is not None:
            continue

        # Determine the date columns range (C to last_col-1)
        first_date_col = 'C'
        last_date_col = get_column_letter(last_col - 1)

        # Create the sum formula
        formula = f"=SUM({first_date_col}{row_idx}:{last_date_col}{row_idx})"
        ws[f"{total_col_letter}{row_idx}"] = formula

        # For debugging, print sample rows
        if row_idx <= 5 or row_idx >= ws.max_row - 5:
            logger.info(f"  Row {row_idx} ({row[0].value}): {formula}")
        elif row_idx == 6:
            logger.info("    ...")

    logger.info("✓ Total column calculation complete for all relevant rows")


def apply_styling(ws):
    """Apply all styling to the worksheet after calculations"""
    logger.info("🎨 Applying styling to worksheet...")

    # Define styles
    weekend_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    total_row_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    occupancy_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    empty_row_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply grid borders to all cells
    logger.info("  Applying grid borders to all cells...")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Apply styling based on cell content
    logger.info("  Applying conditional formatting...")
    for row in ws.iter_rows():
        # Style empty rows with black fill
        if all(cell.value is None for cell in row):
            for cell in row:
                cell.fill = empty_row_fill
            continue

        # Check for date headers (they start with day abbreviations)
        for cell in row:
            if isinstance(cell.value, str) and cell.value.split()[0] in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat',
                                                                         'Sun']:
                if cell.value.startswith(('Fri', 'Sat', 'Sun')):
                    cell.fill = weekend_fill

        # Style Total rows
        if row[0].value and 'Total' in str(row[0].value):
            for cell in row:
                cell.fill = total_row_fill
                cell.font = bold_font

        # Style Occupancy (Πληρότητα) rows
        if row[0].value == 'Πληρότητα':
            for cell in row:
                cell.fill = occupancy_fill

    # Adjust column widths
    logger.info("  Adjusting column widths...")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    logger.info("✓ Styling applied successfully")


def process_stage7(input_file, output_file):
    """Main processing function for stage 7"""
    logger.info("🚀 Starting Stage 7 processing...")

    # Load workbook
    try:
        wb = load_workbook(input_file)
        ws = wb['Stage4 Results']
    except Exception as e:
        logger.info(f"❌ Error loading workbook: {e}")
        return

    # Step 1: Detect breakpoints
    breakpoints = detect_breakpoints(ws)

    # Step 2: Locate target rows
    target_data = locate_target_rows(ws)

    # Step 3: Process each category (sums and capacities)
    logger.info("📊 Calculating sums and capacities...")
    for category_name, data in target_data.items():
        if data['sum_row']:  # Only process main categories
            process_category(ws, category_name, data, breakpoints)

    # Step 4: Calculate occupancy rates
    calculate_occupancy_rates(ws, target_data)

    # Step 5: Calculate Total column for ALL relevant rows
    calculate_total_column(ws)

    # Step 6: Apply all styling
    apply_styling(ws)

    # Save results
    logger.info("💾 Saving results...")
    try:
        wb.save(output_file)
        logger.info(f"✅ Success! Output saved to {output_file}")
        logger.info("🔹 Includes: All calculations with complete styling")
    except Exception as e:
        logger.info(f"❌ Error saving workbook: {e}")


if __name__ == "__main__":
    # Configuration
    INPUT_FILE = 'per_zone_stage6_output.xlsx'
    OUTPUT_FILE = 'per_zone_stage7_output.xlsx'

    # Run the processing
    process_stage7(INPUT_FILE, OUTPUT_FILE)
