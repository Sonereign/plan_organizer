from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from logger import logger

# Constants
GREEK_DAYS = {
    "Mon": "Δευ", "Tue": "Τρι", "Wed": "Τετ", "Thu": "Πεμ",
    "Fri": "Παρ", "Sat": "Σαβ", "Sun": "Κυρ"
}

DAY_COLORS = {
    "Παρ": "ADD8E6",  # Light Blue (Friday)
    "Σαβ": "90EE90",  # Light Green (Saturday)
    "Κυρ": "FFB6C1"  # Light Pink (Sunday)
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep"]
USE_FORMULAS = True
DO_CALCULATIONS = True


def load_and_prepare_data(input_file):
    """Load data from Excel and prepare it for processing."""
    df = pd.read_excel(input_file, header=None)
    headers = df.iloc[0]
    df = df[1:].reset_index(drop=True)
    df.columns = headers
    if "Capacity" in df.columns:
        df = df.drop(columns=["Capacity"])
    return df, headers


def format_date_column(col):
    """Format a single date column into Greek day and date."""
    try:
        date = pd.to_datetime(col, dayfirst=True, errors="coerce")
        if pd.notna(date):
            greek_day = GREEK_DAYS.get(date.strftime("%a"), date.strftime("%a"))
            return f"{greek_day} {date.strftime('%d/%m')}"
        else:
            return col
    except Exception:
        return col


def format_dates(df):
    """Format all date columns in the DataFrame."""
    date_columns = df.columns[1:]
    formatted_columns = [df.columns[0]] + [format_date_column(col) for col in date_columns]
    df.columns = formatted_columns
    return df


def find_camping_first_index(df):
    """Find the first row index where 'Camping' appears in the first column."""
    return df[df.iloc[:, 0].astype(str).str.startswith("Camping")].index.min()


def insert_totals_and_spacing(df, split_index):
    """Insert 'Total Rooms' before camping section and 'Total Camping' at the end, with one empty row in between."""
    total_rooms_row = pd.DataFrame([["Total Rooms"] + [""] * (len(df.columns) - 1)], columns=df.columns)
    total_camping_row = pd.DataFrame([["Total Camping"] + [""] * (len(df.columns) - 1)], columns=df.columns)
    empty_row = pd.DataFrame([[""] * len(df.columns)], columns=df.columns)

    # Split the dataframe into room and camping sections
    top_part = df.iloc[:split_index]
    bottom_part = df.iloc[split_index:]

    # Concatenate parts, ensuring correct order and no extra empty row at the end
    df = pd.concat([top_part, total_rooms_row, empty_row, bottom_part, total_camping_row], ignore_index=True)

    return df


def apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col):
    if not USE_FORMULAS:
        return apply_column_sums_noform(ws, total_rooms_row, total_camping_row, max_col)

    if DO_CALCULATIONS:
        """Apply Excel formulas to calculate column sums."""
        for col in range(2, max_col + 2):
            col_letter = ws.cell(row=1, column=col).column_letter
            if total_rooms_row:
                ws.cell(row=total_rooms_row, column=col).value = f"=SUM({col_letter}2:{col_letter}{total_rooms_row - 1})"
            if total_camping_row:
                ws.cell(row=total_camping_row, column=col).value = f"=SUM({col_letter}{total_rooms_row + 2}:{col_letter}{total_camping_row - 2})"


def apply_column_sums_noform(ws, total_rooms_row, total_camping_row, max_col):
    """Directly calculate and insert column sums without using Excel formulas."""
    if DO_CALCULATIONS:
        for col in range(2, max_col + 2):
            column_sum_rooms = 0
            column_sum_camping = 0

            # Calculate sum for rooms (from row 2 to the row before 'Total Rooms')
            if total_rooms_row:
                for row in range(2, total_rooms_row):
                    value = ws.cell(row=row, column=col).value
                    if value is None:
                        continue
                    if isinstance(value, (int, float)):  # Ensure it's a number
                        column_sum_rooms += value
                ws.cell(row=total_rooms_row, column=col).value = column_sum_rooms

            # Calculate sum for camping (from the row after 'Total Rooms' to the row before 'Total Camping')
            if total_camping_row:
                for row in range(total_rooms_row + 2, total_camping_row):
                    value = ws.cell(row=row, column=col).value
                    if value is None:
                        continue
                    if isinstance(value, (int, float)):  # Ensure it's a number
                        column_sum_camping += value
                ws.cell(row=total_camping_row, column=col).value = column_sum_camping



def apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row):
    """Apply Excel formulas to calculate row sums and percentages."""
    total_column = max_col + 1
    # separator_column_1 = total_column + 1  # First separator (before "Percent to Total")
    # percent_column = separator_column_1 + 1  # "Percent to Total" column
    percent_column = total_column + 1  # "Percent to Total" column
    # separator_column_2 = percent_column + 1  # Second separator (after "Percent to Total")
    current_year = datetime.now().year

    # Add the "Total" column
    add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year)

    # Add the first separator column
    # add_separator_column(ws, max_row, separator_column_1)

    # Add the "Percent to Total" column
    add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, current_year)

    # Add the second separator column
    # add_separator_column(ws, max_row, separator_column_2)

    # Add monthly sums after the second separator
    # add_monthly_sums(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row)
    add_monthly_sums(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row)


def add_total_column(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year):
    if not USE_FORMULAS:
        return add_total_column_direct_noform(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year)

    """Add a 'Total' column to calculate row sums."""
    ws.cell(row=1, column=total_column).value = f"Total {current_year}"
    ws.cell(row=1, column=total_column).font = Font(bold=True)
    if DO_CALCULATIONS:
        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                first_col_letter = ws.cell(row=row, column=2).column_letter
                last_col_letter = ws.cell(row=row, column=max_col).column_letter
                ws.cell(row=row, column=total_column).value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                ws.cell(row=row, column=total_column).fill = YELLOW_FILL
                ws.cell(row=row, column=total_column).font = Font(bold=True)


def add_total_column_direct_noform(ws, max_row, max_col, total_column, total_rooms_row, total_camping_row, current_year):
    """Directly calculate and insert row sums into the 'Total' column without using Excel formulas."""
    ws.cell(row=1, column=total_column).value = f"Total {current_year}"
    ws.cell(row=1, column=total_column).font = Font(bold=True)

    if DO_CALCULATIONS:
        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                row_sum = 0
                for col in range(2, max_col + 1):  # Iterate through the data columns
                    value = ws.cell(row=row, column=col).value
                    if value is None:
                        continue
                    if isinstance(value, (int, float)):  # Ensure the value is a number
                        row_sum += value
                ws.cell(row=row, column=total_column).value = row_sum
                ws.cell(row=row, column=total_column).fill = YELLOW_FILL
                ws.cell(row=row, column=total_column).font = Font(bold=True)


def add_percentage_column(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, current_year):
    if not USE_FORMULAS:
        return add_percentage_column_direct_noform(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, current_year)

    """Add a 'Percent to Total' column to calculate percentages."""
    ws.cell(row=1, column=percent_column).value = f"Percent to Total {current_year}"
    ws.cell(row=1, column=percent_column).font = Font(bold=True)
    if DO_CALCULATIONS:
        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                total_rooms_col_letter = ws.cell(row=total_rooms_row, column=total_column).column_letter
                total_camping_col_letter = ws.cell(row=total_camping_row, column=total_column).column_letter
                current_row_col_letter = ws.cell(row=row, column=total_column).column_letter

                if row < total_rooms_row:  # Rooms section
                    ws.cell(row=row, column=percent_column).value = f"=IF({total_rooms_col_letter}{total_rooms_row}<>0, {current_row_col_letter}{row}/{total_rooms_col_letter}{total_rooms_row}, 0)"
                elif row > total_rooms_row + 1:  # Camping section
                    ws.cell(row=row, column=percent_column).value = f"=IF({total_camping_col_letter}{total_camping_row}<>0, {current_row_col_letter}{row}/{total_camping_col_letter}{total_camping_row}, 0)"

                ws.cell(row=row, column=percent_column).number_format = "0.00%"

    ws.column_dimensions[ws.cell(row=1, column=percent_column).column_letter].width = 15


def add_percentage_column_direct_noform(ws, max_row, total_column, percent_column, total_rooms_row, total_camping_row, current_year):
    """Add a 'Percent to Total' column and calculate percentages directly."""
    ws.cell(row=1, column=percent_column).value = f"Percent to Total {current_year}"
    ws.cell(row=1, column=percent_column).font = Font(bold=True)

    if DO_CALCULATIONS:
        # Manually compute total values for rooms and camping (summing up relevant rows)
        total_rooms_value = 0
        total_camping_value = 0

        # Sum the values for rooms and camping based on the rows you want to sum
        for row in range(2, total_rooms_row):  # Adjust range for rooms section
            current_value = ws.cell(row=row, column=total_column).value
            if isinstance(current_value, (int, float)):  # Ensure it's numeric
                total_rooms_value += current_value

        for row in range(total_rooms_row + 2, max_row + 1):  # Adjust range for camping section
            current_value = ws.cell(row=row, column=total_column).value
            if isinstance(current_value, (int, float)):  # Ensure it's numeric
                total_camping_value += current_value

        for row in range(2, max_row + 1):
            if row not in [total_rooms_row, total_camping_row]:
                current_value = ws.cell(row=row, column=total_column).value
                if isinstance(current_value, (int, float)):  # Ensure the current value is numeric
                    if row < total_rooms_row:  # Rooms section
                        if total_rooms_value and total_rooms_value != 0:  # Avoid division by zero
                            percentage = current_value / total_rooms_value
                            ws.cell(row=row, column=percent_column).value = percentage
                        else:
                            ws.cell(row=row, column=percent_column).value = 0
                    elif row > total_rooms_row + 1:  # Camping section
                        if total_camping_value and total_camping_value != 0:  # Avoid division by zero
                            percentage = current_value / total_camping_value
                            ws.cell(row=row, column=percent_column).value = percentage
                        else:
                            ws.cell(row=row, column=percent_column).value = 0

                    # Format as percentage
                    ws.cell(row=row, column=percent_column).number_format = "0.00%"

    ws.column_dimensions[ws.cell(row=1, column=percent_column).column_letter].width = 15



def add_monthly_sums(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row):
    if not USE_FORMULAS:
        return add_monthly_sums_direct_noform(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row)

    """Add monthly sum columns and calculate their sums."""
    month_ranges = find_monthly_column_ranges(ws, total_column)
    month_start_col = separator_column_2 + 1  # Start after the second separator
    current_year = datetime.now().year

    for i, month in enumerate(MONTHS):
        month_col = month_start_col + i
        ws.cell(row=1, column=month_col).value = f"{month} {current_year}"
        ws.cell(row=1, column=month_col).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=month_col).column_letter].width = 12

        if month in month_ranges:
            first_col_letter = ws.cell(row=1, column=month_ranges[month][0]).column_letter
            last_col_letter = ws.cell(row=1, column=month_ranges[month][1]).column_letter
            if DO_CALCULATIONS:
                for row in range(2, max_row + 1):
                    if row not in [total_rooms_row, total_camping_row]:
                        ws.cell(row=row, column=month_col).value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"

    # Add separator column after the last month column
    # add_separator_column(ws, max_row, month_start_col + len(MONTHS))


def add_monthly_sums_direct_noform(ws, max_row, total_column, separator_column_2, total_rooms_row, total_camping_row):
    """Directly calculate monthly sums and insert them into the columns."""
    month_ranges = find_monthly_column_ranges(ws, total_column)
    month_start_col = separator_column_2 + 1  # Start after the second separator
    current_year = datetime.now().year

    for i, month in enumerate(MONTHS):
        month_col = month_start_col + i
        ws.cell(row=1, column=month_col).value = f"{month} {current_year}"
        ws.cell(row=1, column=month_col).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=month_col).column_letter].width = 12

        if DO_CALCULATIONS:
            if month in month_ranges:
                first_col_index = month_ranges[month][0] - 1  # Adjust for 0-based indexing
                last_col_index = month_ranges[month][1] - 1  # Adjust for 0-based indexing

                for row in range(2, max_row + 1):
                    if row not in [total_rooms_row, total_camping_row]:

                        # Sum the values in the row directly for the given month range
                        row_sum = 0

                        for col in range(first_col_index, last_col_index + 1):
                            cell_value = ws.cell(row=row, column=col + 1).value  # Adjust for 1-based indexing
                            if isinstance(cell_value, (int, float)):  # Check if the value is numeric
                                row_sum += cell_value

                        ws.cell(row=row, column=month_col).value = row_sum

    # Add separator column after the last month column
    # add_separator_column(ws, max_row, month_start_col + len(MONTHS))


def find_monthly_column_ranges(ws, total_column):
    """Find the first and last column for each month."""
    month_ranges = {}
    for col in range(2, total_column):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            for i, month in enumerate(MONTHS):
                month_number = str(i + 4).zfill(2)  # "Apr" = 04, "May" = 05, etc.
                if header_value.endswith(f'/{month_number}'):
                    if month not in month_ranges:
                        month_ranges[month] = [col, col]
                    else:
                        month_ranges[month][1] = col
    return month_ranges


def add_separator_column(ws, max_row, separator_column):
    """Add a black-filled separator column."""
    for row in range(1, max_row + 1):
        separator_cell = ws.cell(row=row, column=separator_column)
        separator_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.column_dimensions[ws.cell(row=1, column=separator_column).column_letter].width = 3


def apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row):
    """Apply formatting to the worksheet."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for col_num in range(2, max_col + 1):
        column_title = ws.cell(row=1, column=col_num).value
        if column_title:
            day_part = column_title.split(" ")[0]
            if day_part in DAY_COLORS:
                ws.cell(row=1, column=col_num).fill = PatternFill(
                    start_color=DAY_COLORS[day_part], end_color=DAY_COLORS[day_part], fill_type="solid"
                )

    for row in [total_rooms_row, total_camping_row]:
        if row:
            for col in range(1, max_col + 2):
                ws.cell(row=row, column=col).fill = YELLOW_FILL
                ws.cell(row=row, column=col).font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col + 1):
        for cell in row:
            cell.border = THIN_BORDER

    ws.freeze_panes = "B2"


def apply_excel_formatting_and_formulas(output_file):
    """Apply formatting and formulas to the output Excel file."""
    wb = load_workbook(output_file)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    total_rooms_row = total_camping_row = None

    for row in range(2, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == "Total Rooms":
            total_rooms_row = row
        elif cell_value == "Total Camping":
            total_camping_row = row

    apply_column_sum_formulas(ws, total_rooms_row, total_camping_row, max_col)
    apply_row_sum_formulas(ws, max_row, max_col, total_rooms_row, total_camping_row)
    apply_formatting(ws, max_col, max_row, total_rooms_row, total_camping_row)

    wb.save(output_file)


def per_nat_stage1_finalizer(input_file, output_file):
    """Process reservations and generate the output Excel file."""
    logger.info("#######################################################")
    logger.info(f"Running Per Nationality Stage 1 with {input_file=} .....")
    df, headers = load_and_prepare_data(input_file)
    df = format_dates(df)
    split_index = find_camping_first_index(df)
    df = insert_totals_and_spacing(df, split_index)
    df.to_excel(output_file, index=False, engine='openpyxl')
    apply_excel_formatting_and_formulas(output_file)
    logger.info(f"Per Nationality Stage 1 completed. File saved as {output_file}")


if __name__ == "__main__":
    # Default file paths (for standalone execution)
    INPUT_FILE = "./sources/availabilityPerNationality2025.xls"
    OUTPUT_FILE = "per_nat_stage1_finalizer_output.xlsx"

    # Run stage5
    per_nat_stage1_finalizer(INPUT_FILE, OUTPUT_FILE)